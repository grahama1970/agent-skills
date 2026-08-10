"""CLI for /extract-entities skill.

Thin httpx client for the memory daemon's /extract-entities endpoint.
No graph_memory imports — all AQL lives in the memory project.

Inputs: Question text via CLI arg or --text flag.
Outputs: JSON EntityExtractionResult to stdout.
Failures: Falls back to empty result if daemon unavailable.
"""

from __future__ import annotations

import json
import os
import re
import sys
from functools import lru_cache
from pathlib import Path

import typer
from loguru import logger
from rich.console import Console
from rich.table import Table

# All commands use httpx to call the memory daemon — no graph_memory imports.

app = typer.Typer(
    name="extract-entities",
    help="Extract control IDs, phrases, and relationships from question text.",
    no_args_is_help=True,
)
console = Console()

_DEFAULT_SPARTA_SOURCE_WORKBOOK = Path(
    "${HOME}/workspace/experiments/sparta/data/source/SPARTA-Data.xlsx"
)

_SPARTA_CONTROL_RETURN_FIELDS = [
    "_id",
    "_key",
    "control_id",
    "id",
    "name",
    "source_framework",
    "framework",
    "node_type",
    "control_type",
    "description",
    "definition",
    "short_description",
    "category",
    "source_category",
    "domain",
    "cluster",
    "namespace",
    "aliases",
    "source_workbook",
    "source_workbook_sheet",
    "source_workbook_row",
    "label",
    "title",
]

_MEMORY_LIST_MAX_LIMIT = 500

_NAME_LOOKUP_STOP_WORDS = {
    "a",
    "an",
    "and",
    "about",
    "are",
    "do",
    "does",
    "for",
    "in",
    "is",
    "know",
    "me",
    "of",
    "on",
    "tell",
    "the",
    "to",
    "what",
    "which",
    "you",
}


def _make_memory_client():
    import httpx

    memory_base_url = os.getenv("MEMORY_API_BASE", "").strip()
    if memory_base_url:
        return httpx.Client(base_url=memory_base_url, timeout=15)

    socket_path = os.getenv("MEMORY_SOCKET", "/run/user/1000/embry/memory.sock")
    transport = httpx.HTTPTransport(uds=socket_path)
    return httpx.Client(transport=transport, base_url="http://localhost", timeout=15)


def _split_tokens(text: str, delimiter: str) -> list[str]:
    mode = delimiter.strip()
    if mode.lower() == "auto":
        return [token for token in re.split(r"[,\s;]+", text) if token]
    if mode.lower() == "nlp":
        return []
    return [token.strip() for token in text.split(mode) if token.strip()]


_KNOWN_INTERFACE_SKILLS = {
    "analytics",
    "ask",
    "assess",
    "create-evidence-case",
    "create-figure",
    "create-table",
    "extract-entities",
    "memory",
}

_SAMPLE_DATA_TERMS = ("sample", "example", "demo", "mock", "placeholder", "fictional")
_DESCRIPTOR_SUBTERMS_TO_SUPPRESS = {"link"}

_DATA_FILE_SUFFIXES = (
    ".arrow",
    ".csv",
    ".json",
    ".jsonl",
    ".md",
    ".parquet",
    ".tsv",
    ".txt",
)


def _span_for(text: str, mention: str) -> list[int]:
    start = text.lower().find(mention.lower())
    if start < 0:
        return [0, 0]
    return [start, start + len(mention)]


def _first_text(doc: dict, *fields: str) -> str:
    for field in fields:
        value = doc.get(field)
        if isinstance(value, str) and value.strip():
            return value
    return ""


def _normalize_sparta_control_doc(doc: dict) -> dict:
    """Normalize a sparta_controls document into the extractor entity shape."""
    control_id = str(doc.get("control_id") or doc.get("id") or "").strip()
    name = str(doc.get("name") or doc.get("title") or control_id).strip()
    framework = str(doc.get("source_framework") or doc.get("framework") or "").strip()
    control_type = str(doc.get("control_type") or doc.get("node_type") or "").strip()
    description = _first_text(doc, "description", "definition", "short_description")
    short_description = _first_text(doc, "short_description") or description[:280]
    category = _first_text(doc, "category", "source_category")
    aliases = doc.get("aliases") if isinstance(doc.get("aliases"), list) else []

    normalized = {
        "_id": doc.get("_id", ""),
        "_key": doc.get("_key", ""),
        "id": doc.get("_id") or control_id,
        "canonical_id": control_id,
        "control_id": control_id,
        "label": control_id,
        "canonical_name": name,
        "name": name,
        "framework": framework,
        "source_framework": framework,
        "entity_type": control_type or "control",
        "type": control_type or "control",
        "control_type": control_type,
        "description": description,
        "definition": description,
        "short_description": short_description,
        "category": category,
        "domain": doc.get("domain", ""),
        "cluster": doc.get("cluster", ""),
        "namespace": doc.get("namespace", ""),
        "aliases": aliases,
        "source_workbook": doc.get("source_workbook", ""),
        "source_workbook_sheet": doc.get("source_workbook_sheet", ""),
        "source_workbook_row": doc.get("source_workbook_row", ""),
        "exists": True,
    }
    return {key: value for key, value in normalized.items() if value not in ("", None, [])}


def _list_documents_paged(
    *,
    client,
    collection: str,
    limit: int,
    return_fields: list[str],
    q: str | None = None,
    filters: dict | None = None,
) -> list[dict]:
    """Read Memory /list in bounded pages.

    The daemon caps each /list request at 500 rows. A caller-provided larger
    limit is a total cap, not a single request size.
    """
    total_limit = max(0, int(limit))
    if total_limit == 0:
        return []

    def _page(offset: int, batch_limit: int) -> list[dict]:
        payload = {
            "collection": collection,
            "limit": batch_limit,
            "offset": offset,
            "return_fields": return_fields,
        }
        if q:
            payload["q"] = q
        if filters:
            payload["filters"] = filters
        r = client.post("/list", json=payload)
        r.raise_for_status()
        return r.json().get("documents", [])

    # The daemon caps a page at 500 rows, so a large dictionary needs many
    # requests -- 782 of them for 391k entities. Issued sequentially that is ~20s
    # of pure round-trip latency for work the server answers in 25ms a page.
    # Fetched concurrently instead: the cap is the daemon's, the waiting was ours.
    #
    # Probing the first page tells us whether concurrency is worth it at all;
    # small dictionaries stay on the simple path.
    first = _page(0, min(_MEMORY_LIST_MAX_LIMIT, total_limit))
    docs: list[dict] = list(first)
    if len(first) >= _MEMORY_LIST_MAX_LIMIT and total_limit > _MEMORY_LIST_MAX_LIMIT:
        from concurrent.futures import ThreadPoolExecutor

        offsets = list(range(_MEMORY_LIST_MAX_LIMIT, total_limit, _MEMORY_LIST_MAX_LIMIT))
        with ThreadPoolExecutor(max_workers=32) as pool:
            futures = {
                offset: pool.submit(_page, offset,
                                    min(_MEMORY_LIST_MAX_LIMIT, total_limit - offset))
                for offset in offsets
            }
            # Ordered by offset so the dictionary is deterministic across runs.
            for offset in offsets:
                batch = futures[offset].result()
                if not batch:
                    break
                docs.extend(batch)
        offset += len(batch)

    return docs


def _source_workbook_path() -> Path:
    return Path(os.getenv("SPARTA_SOURCE_WORKBOOK", str(_DEFAULT_SPARTA_SOURCE_WORKBOOK)))


@lru_cache(maxsize=1)
def _load_source_workbook_controls() -> tuple[dict[str, dict], dict[str, list[dict]]]:
    """Load authoritative SPARTA countermeasure rows as a stale-corpus fallback.

    ArangoDB remains the normal vocabulary source. This path is only used when
    `/memory` misses an exact control/category/name that exists in the local
    source workbook.
    """
    path = _source_workbook_path()
    if not path.exists():
        return {}, {}
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        logger.warning(f"openpyxl unavailable for SPARTA workbook fallback: {exc}")
        return {}, {}

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["SPARTA Countermeasures"]
    except Exception as exc:
        logger.warning(f"failed loading SPARTA workbook fallback {path}: {exc}")
        return {}, {}

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, {}
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    indices = {header: index for index, header in enumerate(headers)}

    def value(row: tuple, field: str) -> str:
        index = indices.get(field)
        if index is None or index >= len(row):
            return ""
        raw = row[index]
        if raw is None:
            return ""
        text_value = str(raw).strip()
        return "" if text_value.lower() == "nan" else text_value

    def source_list(raw: str) -> list[str]:
        if not raw:
            return []
        items: list[str] = []
        for item in raw.replace("\n", ",").split(","):
            cleaned = item.strip()
            if cleaned and cleaned.lower() not in {"none", "nan"}:
                items.append(cleaned)
        return items

    by_control_id: dict[str, dict] = {}
    by_term: dict[str, list[dict]] = {}
    for row_number, row in enumerate(rows[1:], start=2):
        control_id = value(row, "ID")
        if not control_id:
            continue
        category = value(row, "Category")
        name = value(row, "Name")
        description = value(row, "Description")
        sources = value(row, "Sources")
        doc = {
            "_id": f"sparta_source_workbook/{control_id}",
            "_key": f"source__{control_id}",
            "control_id": control_id,
            "id": control_id,
            "name": name,
            "category": category,
            "source_category": category,
            "source_framework": "SPARTA",
            "framework": "SPARTA",
            "control_type": "countermeasure",
            "description": description,
            "definition": description,
            "short_description": description[:280] if description else "",
            "sources": source_list(sources),
            "source_url": sources,
            "aliases": [term for term in (category,) if term],
            "source": "sparta_source_workbook_fallback",
            "source_workbook": str(path),
            "source_workbook_sheet": "SPARTA Countermeasures",
            "source_workbook_row": row_number,
        }
        by_control_id[control_id.upper()] = doc
        for term in (control_id, name, category):
            cleaned = str(term or "").strip().lower()
            if cleaned:
                by_term.setdefault(cleaned, []).append(doc)
    return by_control_id, by_term


def _lookup_source_workbook_control_doc(
    *,
    control_id: str | None = None,
    name: str | None = None,
) -> dict | None:
    by_control_id, by_term = _load_source_workbook_controls()
    if control_id:
        doc = by_control_id.get(control_id.strip().upper())
        if doc:
            return dict(doc)
    if name:
        docs = by_term.get(name.strip().lower(), [])
        if len(docs) == 1:
            return dict(docs[0])
    return None


def _lookup_sparta_control_doc(
    client,
    *,
    control_id: str | None = None,
    name: str | None = None,
) -> dict | None:
    """Return one exact sparta_controls document by control_id or name."""
    filters: dict[str, str] = {}
    if control_id:
        filters["control_id"] = control_id
    elif name:
        filters["name"] = name
    else:
        return None

    try:
        r = client.post("/list", json={
            "collection": "sparta_controls",
            "limit": 1,
            "filters": filters,
            "return_fields": _SPARTA_CONTROL_RETURN_FIELDS,
        })
        docs = r.json().get("documents", [])
    except Exception as e:
        logger.warning(f"sparta_controls lookup failed for {filters}: {e}")
        docs = []
    if docs:
        return docs[0]
    return _lookup_source_workbook_control_doc(control_id=control_id, name=name)


def _lookup_sparta_control_doc_by_name(client, mention: str) -> dict | None:
    doc = _lookup_sparta_control_doc(client, name=mention)
    if doc:
        return doc
    upper = mention.upper()
    if upper != mention:
        return _lookup_sparta_control_doc(client, name=upper)
    return None


def _candidate_control_name_mentions(text: str) -> list[str]:
    tokens = [
        token.strip(".,;:!?()[]{}<>\"'")
        for token in text.split()
        if token.strip(".,;:!?()[]{}<>\"'")
    ]
    candidates: list[str] = []
    seen: set[str] = set()
    for phrase in tokens:
        lower = phrase.lower()
        if lower in _NAME_LOOKUP_STOP_WORDS:
            continue
        if len(phrase) < 4:
            continue
        if not (phrase.isupper() or any(char.isdigit() for char in phrase)):
            continue
        if lower in seen:
            continue
        seen.add(lower)
        candidates.append(phrase)
    return candidates


def _control_metadata_from_entity(entity: dict) -> dict:
    return {
        "_id": entity.get("_id", ""),
        "_key": entity.get("_key", ""),
        "control_id": entity.get("control_id") or entity.get("canonical_id") or entity.get("label", ""),
        "name": entity.get("name") or entity.get("canonical_name", ""),
        "framework": entity.get("source_framework") or entity.get("framework", ""),
        "domain": entity.get("domain", ""),
        "type": entity.get("control_type") or entity.get("entity_type") or entity.get("type", ""),
        "description": entity.get("description", ""),
        "definition": entity.get("definition", ""),
        "short_description": entity.get("short_description", ""),
        "category": entity.get("category") or entity.get("source_category", ""),
        "aliases": entity.get("aliases", []),
    }


def _unsupported_term_id(mention: str) -> str:
    slug = "".join(char.lower() if char.isalnum() else "_" for char in mention).strip("_")
    while "__" in slug:
        slug = slug.replace("__", "_")
    return f"unsupported:{slug or 'term'}"


def _term_node_id(mention: str, prefix: str = "term") -> str:
    slug = "".join(char.lower() if char.isalnum() else "_" for char in mention).strip("_")
    while "__" in slug:
        slug = slug.replace("__", "_")
    return f"{prefix}:{slug or 'term'}"


def _glossary_from_entities(
    *,
    resolved: list[dict],
    descriptors: list[dict] | None = None,
    external: list[dict],
    unresolved: list[dict],
) -> list[dict]:
    """Collate extraction output into a UI/evidence friendly glossary.

    The extractor owns this normalization so downstream skills can snapshot the
    same terms instead of independently rebuilding glossary semantics.
    """
    entries: list[dict] = []
    seen: set[tuple[str, str, tuple]] = set()

    def add(entry: dict) -> None:
        identifier = str(entry.get("id") or entry.get("name") or entry.get("mention") or "").strip().lower()
        span = entry.get("span")
        span_key = tuple(span) if isinstance(span, list) else ()
        state = "grounded" if entry.get("grounded") else "unsupported"
        key = (identifier, state, span_key)
        if not identifier or key in seen:
            return
        seen.add(key)
        entries.append({k: v for k, v in entry.items() if v not in ("", None, [])})

    for ent in resolved:
        control_id = str(ent.get("control_id") or ent.get("canonical_id") or ent.get("label") or "").strip()
        name = str(ent.get("name") or ent.get("canonical_name") or control_id).strip()
        add({
            "id": control_id or ent.get("id"),
            "control_id": control_id,
            "name": name,
            "mention": ent.get("mention") or control_id or name,
            "framework": ent.get("source_framework") or ent.get("framework"),
            "type": ent.get("control_type") or ent.get("entity_type") or ent.get("type"),
            "description": ent.get("description"),
            "definition": ent.get("definition"),
            "short_description": ent.get("short_description"),
            "category": ent.get("category") or ent.get("source_category"),
            "aliases": ent.get("aliases"),
            "span": ent.get("span"),
            "grounded": True,
            "exists": True,
            "source": ent.get("source") or "sparta_controls",
        })

    for ent in descriptors or []:
        mention = str(ent.get("mention") or ent.get("name") or "").strip()
        add({
            "id": ent.get("id") or _term_node_id(mention, "descriptor"),
            "name": mention,
            "mention": mention,
            "framework": ent.get("framework"),
            "type": ent.get("entity_type") or "control_descriptor",
            "span": ent.get("span"),
            "grounded": True,
            "exists": True,
            "source": ent.get("source") or "sparta_controls_parenthetical_guard",
            "control_id": ent.get("control_id"),
            "descriptor_kind": ent.get("descriptor_kind"),
            "matched_fields": ent.get("matched_fields"),
            "canonical_name": ent.get("canonical_name"),
            "category": ent.get("category"),
        })

    for ent in external:
        mention = str(ent.get("mention") or ent.get("normalized_text") or "").strip()
        add({
            "id": ent.get("id") or _unsupported_term_id(mention),
            "name": mention,
            "mention": mention,
            "framework": ent.get("framework"),
            "type": ent.get("entity_type") or "external_entity",
            "span": ent.get("span"),
            "grounded": False,
            "exists": False,
            "reason": ent.get("routing_effect") or ent.get("reason") or "not_grounded_to_sparta_controls",
            "source": ent.get("source") or "external_entity",
            "claimed_against_control_id": ent.get("claimed_against_control_id") or ent.get("control_id"),
            "expected_control_name": ent.get("expected_control_name") or ent.get("canonical_name"),
        })

    for ent in unresolved:
        mention = str(ent.get("mention") or ent.get("term") or ent.get("normalized_text") or "").strip()
        add({
            "id": None,
            "name": mention,
            "mention": mention,
            "framework": ent.get("framework"),
            "type": ent.get("entity_type") or ent.get("type") or "unresolved",
            "span": ent.get("span"),
            "grounded": False,
            "exists": False,
            "reason": ent.get("reason") or "unresolved",
            "source": ent.get("source") or "unresolved_entities",
        })

    return sorted(
        entries,
        key=lambda item: (
            item.get("span", [10**9, 10**9])[0] if isinstance(item.get("span"), list) else 10**9,
            0 if item.get("grounded") else 1,
            str(item.get("name") or item.get("id") or ""),
        ),
    )


def _entity_nodes_from_extraction(
    *,
    resolved: list[dict],
    descriptors: list[dict] | None = None,
    external: list[dict],
    unresolved: list[dict],
    domain_terms: list[dict],
) -> list[dict]:
    nodes: list[dict] = []
    seen: set[tuple[str, tuple]] = set()

    def add(node: dict) -> None:
        node_id = str(node.get("id") or "").strip()
        span = node.get("span")
        span_key = tuple(span) if isinstance(span, list) else ()
        if not node_id:
            return
        key = (node_id, span_key)
        if key in seen:
            return
        seen.add(key)
        nodes.append({k: v for k, v in node.items() if v not in ("", None, [])})

    for ent in resolved:
        cid = str(ent.get("control_id") or ent.get("canonical_id") or ent.get("label") or "").strip()
        name = str(ent.get("name") or ent.get("canonical_name") or cid).strip()
        mention = ent.get("mention") or cid or name
        add({
            "id": cid or ent.get("id"),
            "node_kind": "control",
            "status": "grounded",
            "proof_role": "entity_anchor",
            "source": {
                "collection": "sparta_controls",
                "object_id": ent.get("_id") or (f"sparta_controls/{cid}" if cid else None),
                "workbook": ent.get("source_workbook"),
                "sheet": ent.get("source_workbook_sheet"),
                "row": ent.get("source_workbook_row"),
            },
            "extracted": {
                "text": mention,
                "span": ent.get("span"),
                "source": ent.get("source") or "sparta_controls",
                "kind": "control_id" if mention == cid else "control_name",
            },
            "metadata": {
                "control_id": cid,
                "name": name,
                "framework": ent.get("source_framework") or ent.get("framework"),
                "type": ent.get("control_type") or ent.get("entity_type") or ent.get("type"),
                "category": ent.get("category") or ent.get("source_category"),
                "grounded": True,
                "exists": True,
            },
        })

    for ent in descriptors or []:
        mention = str(ent.get("mention") or ent.get("name") or "").strip()
        add({
            "id": ent.get("id") or _term_node_id(mention, "descriptor"),
            "node_kind": "control_descriptor",
            "status": "grounded",
            "proof_role": "validated_context",
            "extracted": {
                "text": mention,
                "span": ent.get("span"),
                "source": ent.get("source") or "sparta_controls_parenthetical_guard",
                "kind": ent.get("entity_type") or "control_descriptor",
            },
            "relationship": {
                "type": "describes_control",
                "target_id": ent.get("control_id"),
            },
            "metadata": {
                "name": mention,
                "type": ent.get("entity_type") or "control_descriptor",
                "grounded": True,
                "exists": True,
                "control_id": ent.get("control_id"),
                "descriptor_kind": ent.get("descriptor_kind"),
                "matched_fields": ent.get("matched_fields"),
                "canonical_name": ent.get("canonical_name"),
                "category": ent.get("category"),
            },
        })

    for ent in external:
        mention = str(ent.get("mention") or ent.get("normalized_text") or "").strip()
        add({
            "id": ent.get("id") or _unsupported_term_id(mention),
            "node_kind": "unsupported_term",
            "status": "unsupported",
            "proof_role": "none",
            "extracted": {
                "text": mention,
                "span": ent.get("span"),
                "source": ent.get("source") or "external_entity",
                "kind": ent.get("entity_type") or "external_entity",
            },
            "metadata": {
                "name": mention,
                "normalized_text": ent.get("normalized_text"),
                "type": ent.get("entity_type") or "external_entity",
                "grounded": False,
                "exists": False,
                "reason": ent.get("routing_effect") or ent.get("reason") or "not_grounded_to_sparta_controls",
                "claimed_against_control_id": ent.get("claimed_against_control_id") or ent.get("control_id"),
                "expected_control_name": ent.get("expected_control_name") or ent.get("canonical_name"),
            },
        })

    for ent in unresolved:
        mention = str(ent.get("mention") or ent.get("term") or ent.get("normalized_text") or "").strip()
        add({
            "id": ent.get("id") or _term_node_id(mention, "unresolved"),
            "node_kind": "unresolved_term",
            "status": "unresolved",
            "proof_role": "none",
            "extracted": {
                "text": mention,
                "span": ent.get("span"),
                "source": ent.get("source") or "unresolved_entities",
                "kind": ent.get("entity_type") or ent.get("type") or "unresolved",
            },
            "metadata": {
                "name": mention,
                "type": ent.get("entity_type") or ent.get("type") or "unresolved",
                "grounded": False,
                "exists": False,
                "reason": ent.get("reason") or "unresolved",
            },
        })

    for term in domain_terms:
        mention = str(term.get("text") or term.get("mention") or "").strip()
        add({
            "id": term.get("id") or _term_node_id(mention, "domain"),
            "node_kind": "domain_term",
            "status": "extracted",
            "proof_role": "query_context",
            "extracted": {
                "text": mention,
                "span": term.get("span"),
                "source": term.get("source") or "extract_entities_domain_terms",
                "kind": term.get("kind") or "domain_term",
            },
            "metadata": {
                "name": mention,
                "type": term.get("kind") or "domain_term",
                "grounded": True,
                "exists": True,
            },
        })

    return sorted(
        nodes,
        key=lambda item: (
            item.get("extracted", {}).get("span", [10**9, 10**9])[0]
            if isinstance(item.get("extracted", {}).get("span"), list) else 10**9,
            str(item.get("id") or ""),
        ),
    )


def _span_contains(outer: list | None, inner: list | None) -> bool:
    return (
        isinstance(outer, list)
        and isinstance(inner, list)
        and len(outer) == 2
        and len(inner) == 2
        and int(outer[0]) <= int(inner[0])
        and int(inner[1]) <= int(outer[1])
        and outer != inner
    )


def _descriptor_subterm_span(descriptor: dict, subterm: str) -> list[int] | None:
    mention = str(descriptor.get("mention") or descriptor.get("name") or "")
    descriptor_span = descriptor.get("span")
    if (
        not mention
        or not isinstance(descriptor_span, list)
        or len(descriptor_span) != 2
    ):
        return None
    offset = mention.lower().find(subterm.lower())
    if offset < 0:
        return None
    start = int(descriptor_span[0]) + offset
    return [start, start + len(subterm)]


def _split_domain_terms_by_descriptor_coverage(
    *,
    domain_terms: list[dict],
    descriptors: list[dict],
) -> tuple[list[dict], list[dict]]:
    """Suppress context terms already covered by a parent descriptor span."""
    kept: list[dict] = []
    suppressed: list[dict] = []
    suppressed_keys: set[tuple[str, tuple[int, int] | None]] = set()
    for term in domain_terms:
        mention = str(term.get("text") or term.get("mention") or "").strip()
        term_span = term.get("span")
        covering_descriptor = next(
            (
                descriptor for descriptor in descriptors
                if _span_contains(descriptor.get("span"), term_span)
            ),
            None,
        )
        if covering_descriptor:
            grounded = bool(covering_descriptor.get("grounded"))
            span_key = tuple(term_span) if isinstance(term_span, list) and len(term_span) == 2 else None
            suppressed_keys.add((mention.lower(), span_key))
            suppressed.append({
                "id": term.get("id") or _term_node_id(mention, "domain"),
                "node_kind": "domain_term",
                "mention": mention,
                "span": term_span,
                "status": "suppressed",
                "reason": (
                    "covered_by_grounded_control_descriptor"
                    if grounded else "covered_by_unsupported_control_descriptor"
                ),
                "covered_by": covering_descriptor.get("id"),
                "covered_by_mention": covering_descriptor.get("mention") or covering_descriptor.get("name"),
                "covered_by_span": covering_descriptor.get("span"),
            })
        else:
            kept.append(term)
    active_keys = {
        (
            str(term.get("text") or term.get("mention") or "").strip().lower(),
            tuple(term.get("span")) if isinstance(term.get("span"), list) and len(term.get("span")) == 2 else None,
        )
        for term in kept
    }
    for descriptor in descriptors:
        mention = str(descriptor.get("mention") or descriptor.get("name") or "")
        for raw_part in mention.split():
            part = raw_part.strip("()[]{}.,;:!?\"'")
            if part.lower() not in _DESCRIPTOR_SUBTERMS_TO_SUPPRESS:
                continue
            part_span = _descriptor_subterm_span(descriptor, part)
            span_key = tuple(part_span) if part_span else None
            key = (part.lower(), span_key)
            if key in suppressed_keys or key in active_keys:
                continue
            grounded = bool(descriptor.get("grounded"))
            suppressed_keys.add(key)
            suppressed.append({
                "id": _term_node_id(part, "domain"),
                "node_kind": "domain_term",
                "mention": part,
                "span": part_span,
                "status": "suppressed",
                "reason": (
                    "covered_by_grounded_control_descriptor"
                    if grounded else "covered_by_unsupported_control_descriptor"
                ),
                "covered_by": descriptor.get("id"),
                "covered_by_mention": descriptor.get("mention") or descriptor.get("name"),
                "covered_by_span": descriptor.get("span"),
            })
    return kept, suppressed


def _anchor_assertion(anchor: dict) -> dict:
    metadata = anchor.get("metadata", {})
    source = anchor.get("source", {})
    control_id = metadata.get("control_id") or anchor.get("id")
    return {
        "id": f"assert:{str(control_id).lower()}_control_id_resolved",
        "rule_id": "SPARTA_CONTROL_ID_EXACT_MATCH",
        "subject": control_id,
        "predicate": "resolves_to",
        "object": source.get("object_id") or f"sparta_controls/{control_id}",
        "status": "passed",
        "evidence": {
            "control_id": control_id,
            "canonical_name": metadata.get("name"),
            "source_collection": source.get("collection") or "sparta_controls",
            "source_workbook": source.get("workbook"),
            "source_sheet": source.get("sheet"),
            "source_row": source.get("row"),
        },
    }


def _descriptor_assertion(descriptor: dict) -> dict:
    extracted = descriptor.get("extracted", {})
    metadata = descriptor.get("metadata", {})
    relationship = descriptor.get("relationship", {})
    descriptor_id = descriptor.get("id")
    mention = extracted.get("text") or metadata.get("name")
    matched_fields = metadata.get("matched_fields")
    if not matched_fields:
        matched_fields = [metadata.get("descriptor_kind")] if metadata.get("descriptor_kind") else []
    evidence = {
        "mention": mention,
        "span": extracted.get("span"),
        "descriptor_kind": metadata.get("descriptor_kind"),
        "matched_fields": matched_fields,
        "matched_value": mention,
    }
    return {
        "id": f"assert:{str(descriptor_id).replace(':', '_')}_descriptor_validated",
        "rule_id": "PARENTHETICAL_DESCRIPTOR_MATCHES_CONTROL_CATEGORY_OR_ALIAS",
        "subject": descriptor_id,
        "predicate": relationship.get("type") or "describes_control",
        "object": relationship.get("target_id"),
        "status": "passed",
        "evidence": evidence,
    }


def _suppression_assertion(node: dict) -> dict:
    reason = node.get("reason")
    return {
        "id": f"assert:{str(node.get('id')).replace(':', '_')}_suppressed",
        "rule_id": (
            "DOMAIN_TERM_SUBSPAN_COVERED_BY_GROUNDED_DESCRIPTOR"
            if reason == "covered_by_grounded_control_descriptor"
            else "DOMAIN_TERM_SUBSPAN_COVERED_BY_EXTRACTED_PARENT_TERM"
        ),
        "subject": node.get("id"),
        "predicate": "suppressed_because_covered_by",
        "object": node.get("covered_by"),
        "status": "passed",
        "evidence": {
            "mention": node.get("mention"),
            "span": node.get("span"),
            "reason": reason,
            "covering_mention": node.get("covered_by_mention"),
            "covering_span": node.get("covered_by_span"),
        },
    }


def _unsupported_assertion(node: dict) -> dict:
    metadata = node.get("metadata", {})
    return {
        "id": f"assert:{str(node.get('id')).replace(':', '_')}_unsupported",
        "rule_id": "PARENTHETICAL_DESCRIPTOR_MATCHES_CONTROL_CATEGORY_OR_ALIAS",
        "subject": node.get("id"),
        "predicate": "does_not_describe_control",
        "object": metadata.get("claimed_against_control_id"),
        "status": "failed",
        "evidence": {
            "mention": node.get("extracted", {}).get("text") or metadata.get("name"),
            "span": node.get("extracted", {}).get("span"),
            "reason": metadata.get("reason"),
            "expected_control_name": metadata.get("expected_control_name"),
        },
    }


def _proof_packet_from_nodes(
    *,
    query_text: str,
    nodes: list[dict],
    suppressed_nodes: list[dict],
    primary_entity_id: str | None,
) -> dict:
    anchors = [node for node in nodes if node.get("proof_role") == "entity_anchor"]
    validated_context = [node for node in nodes if node.get("proof_role") == "validated_context"]
    context_terms = [node for node in nodes if node.get("proof_role") == "query_context"]
    unsupported = [node for node in nodes if node.get("status") == "unsupported"]
    unsupported_mentions = {
        (tuple(node.get("span") or []), str(node.get("mention") or "").lower())
        for node in unsupported
    }
    unsupported.extend(
        node
        for node in nodes
        if node.get("status") == "unresolved"
        and (tuple(node.get("span") or []), str(node.get("mention") or "").lower())
        not in unsupported_mentions
    )
    assertions = (
        [_anchor_assertion(node) for node in anchors]
        + [_descriptor_assertion(node) for node in validated_context]
        + [_suppression_assertion(node) for node in suppressed_nodes]
        + [_unsupported_assertion(node) for node in unsupported]
    )

    return {
        "generated_by": "deterministic_extractor",
        "llm_used": False,
        "authoritative": True,
        "query_text": query_text,
        "primary_entity_id": primary_entity_id,
        "assertions": assertions,
        "anchors": anchors,
        "validated_context": validated_context,
        "context_terms": context_terms,
        "suppressed_nodes": suppressed_nodes,
        "unsupported_nodes": unsupported,
    }


def _display_summary_from_proof_packet(packet: dict) -> dict:
    bits: list[str] = []
    primary = packet.get("primary_entity_id")
    if primary:
        bits.append(f"{primary} is grounded as the primary extraction anchor.")
    validated = packet.get("validated_context") or []
    if validated:
        mentions = ", ".join(
            str(node.get("extracted", {}).get("text") or node.get("id"))
            for node in validated
        )
        bits.append(f"{mentions} is validated as context attached to the anchor.")
    suppressed = packet.get("suppressed_nodes") or []
    if suppressed:
        mentions = ", ".join(str(node.get("mention") or node.get("id")) for node in suppressed)
        bits.append(f"{mentions} suppressed because it is covered by a grounded descriptor.")
    return {
        "generated_by": "deterministic_template",
        "llm_used": False,
        "authoritative": False,
        "proof_role": "explanation_only",
        "text": " ".join(bits) or "No grounded extraction anchors were found.",
    }


def _agent_contract() -> dict:
    return {
        "use_for_grounding": ["proof_packet.assertions"],
        "authoritative_paths": [
            "proof_packet.assertions",
            "nodes.anchors",
            "nodes.validated_context",
            "nodes.suppressed",
        ],
        "ignore_for_grounding": [
            "display_summary",
            "candidate_nodes",
            "guess_nodes",
            "rationale.summary",
        ],
        "llm_generated_fields_are_never_authoritative": True,
    }


def _node_span(node: dict) -> list[int] | None:
    span = node.get("span") or node.get("extracted", {}).get("span")
    return span if isinstance(span, list) and len(span) == 2 else None


def _node_mention(node: dict) -> str:
    return str(
        node.get("mention")
        or node.get("extracted", {}).get("text")
        or node.get("metadata", {}).get("name")
        or node.get("id")
        or ""
    )


def _compact_node(node: dict) -> dict:
    metadata = dict(node.get("metadata") or {})
    compact = {
        "id": node.get("id"),
        "node_kind": node.get("node_kind"),
        "proof_role": node.get("proof_role"),
        "status": node.get("status"),
        "mention": _node_mention(node),
        "span": _node_span(node),
    }
    keep_metadata = {
        key: metadata.get(key)
        for key in (
            "control_id",
            "name",
            "framework",
            "type",
            "category",
            "aliases",
            "descriptor_kind",
            "matched_fields",
            "canonical_name",
            "grounded",
            "exists",
            "reason",
            "claimed_against_control_id",
            "expected_control_name",
        )
        if metadata.get(key) not in (None, "", [])
    }
    if keep_metadata:
        compact["metadata"] = keep_metadata
    if node.get("relationship"):
        compact["relationship"] = node["relationship"]
    if node.get("reason"):
        compact["reason"] = node.get("reason")
    if node.get("covered_by"):
        compact["covered_by"] = node.get("covered_by")
    return {key: value for key, value in compact.items() if value not in (None, "", [])}


def _nodes_from_proof_packet(packet: dict) -> dict:
    return {
        "anchors": [_compact_node(node) for node in packet.get("anchors", [])],
        "validated_context": [_compact_node(node) for node in packet.get("validated_context", [])],
        "context_terms": [_compact_node(node) for node in packet.get("context_terms", [])],
        "suppressed": [_compact_node(node) for node in packet.get("suppressed_nodes", [])],
        "unsupported": [_compact_node(node) for node in packet.get("unsupported_nodes", [])],
    }


def _compact_candidate_nodes(candidate_nodes: dict, *, primary_entity_id: str | None = None) -> dict:
    primary = str(primary_entity_id or "").strip()
    controls: list[dict] = []
    for node in candidate_nodes.get("controls", []) if isinstance(candidate_nodes, dict) else []:
        cid = str(node.get("control_id") or node.get("id") or "").strip()
        if primary and cid == primary and node.get("match_basis") == "exact_control_id":
            continue
        controls.append({
            key: node.get(key)
            for key in ("id", "control_id", "name", "framework", "type", "match_basis", "reason", "source")
            if node.get(key) not in (None, "", [])
        })

    qras: list[dict] = []
    for node in candidate_nodes.get("qras", []) if isinstance(candidate_nodes, dict) else []:
        qras.append({
            key: node.get(key)
            for key in ("id", "control_id", "question", "match_basis", "source", "review_status")
            if node.get(key) not in (None, "", [])
        })

    return {
        "status": "candidate_only",
        "proof_role": "none",
        "usable_for": ["suggestions", "related_results", "did_you_mean"],
        "not_usable_for": ["grounding", "proof", "approval", "answer_authority", "evidence_gate"],
        "requires_user_confirmation": True,
        "controls": controls,
        "qras": qras,
    }


def _agent_extract_view(full: dict) -> dict:
    packet = full.get("proof_packet") or {}
    primary_entity_id = packet.get("primary_entity_id")
    grounding_ok = bool(full.get("grounding_ok", True))
    decision = dict(full.get("agent_decision") or {})
    decision["primary_entity_id"] = primary_entity_id
    decision["grounding_source"] = "proof_packet.assertions"
    if not grounding_ok:
        decision["safe_to_answer"] = False
        decision["needs_clarification"] = True
        decision["suggested_action"] = "clarify_or_repair"
    proof_packet = {
        "generated_by": packet.get("generated_by"),
        "llm_used": packet.get("llm_used"),
        "authoritative": packet.get("authoritative"),
        "query_text": packet.get("query_text") or full.get("query_text"),
        "primary_entity_id": primary_entity_id,
        "assertions": packet.get("assertions", []),
    }
    counts = {
        "anchors": len(packet.get("anchors", [])),
        "validated_context": len(packet.get("validated_context", [])),
        "context_terms": len(packet.get("context_terms", [])),
        "suppressed": len(packet.get("suppressed_nodes", [])),
        "unsupported": len(packet.get("unsupported_nodes", [])),
        "candidate_controls": len((full.get("candidate_nodes") or {}).get("controls", [])),
        "candidate_qras": len((full.get("candidate_nodes") or {}).get("qras", [])),
    }
    return {
        "view": "agent",
        "ok": full.get("ok", True),
        "grounding_ok": grounding_ok,
        "query_text": full.get("query_text", ""),
        "agent_decision": decision,
        "proof_packet": proof_packet,
        "nodes": _nodes_from_proof_packet(packet),
        "candidate_nodes": _compact_candidate_nodes(full.get("candidate_nodes") or {}, primary_entity_id=primary_entity_id),
        "counts": counts,
        "agent_contract": _agent_contract(),
    }


def _shape_extraction_result(full: dict, view: str) -> dict:
    normalized = (view or "agent").lower()
    if normalized == "legacy":
        shaped = dict(full)
        shaped["view"] = "legacy"
        return shaped
    agent_view = _agent_extract_view(full)
    if normalized == "agent":
        return agent_view
    shaped = dict(full)
    shaped["view"] = normalized
    shaped["agent"] = agent_view
    shaped["nodes"] = agent_view["nodes"]
    if normalized == "debug":
        shaped["debug"] = {
            "response_shape": "verbose_plus_agent_view",
            "proof_assertion_count": len((full.get("proof_packet") or {}).get("assertions", [])),
            "legacy_field_count": len(shaped.keys()),
        }
    return shaped


def _list_collection_docs(client, *, collection: str, filters: dict, limit: int, return_fields: list[str]) -> list[dict]:
    try:
        r = client.post("/list", json={
            "collection": collection,
            "limit": limit,
            "filters": filters,
            "return_fields": return_fields,
        })
        docs = r.json().get("documents", [])
        return [doc for doc in docs if isinstance(doc, dict)]
    except Exception as e:
        logger.warning(f"candidate lookup failed for {collection} {filters}: {e}")
        return []


def _candidate_control_node(doc: dict, *, query: str, match_basis: str, reason: str) -> dict:
    normalized = _normalize_sparta_control_doc(doc)
    return {
        "layer": "sparta_controls",
        "status": "candidate",
        "proof_role": "none",
        "query": query,
        "match_basis": match_basis,
        "reason": reason,
        "id": normalized.get("control_id") or normalized.get("id"),
        "control_id": normalized.get("control_id"),
        "name": normalized.get("name"),
        "framework": normalized.get("framework"),
        "type": normalized.get("type"),
        "description": normalized.get("description"),
        "source": "sparta_controls",
    }


def _candidate_qra_node(doc: dict, *, query: str, control_id: str, collection: str) -> dict:
    return {
        "layer": collection,
        "status": "candidate",
        "proof_role": "none",
        "query": query,
        "match_basis": "same_control_id",
        "reason": f"related QRA candidate for extracted control {control_id}",
        "id": doc.get("qra_id") or doc.get("_id") or doc.get("_key"),
        "_id": doc.get("_id"),
        "_key": doc.get("_key"),
        "control_id": doc.get("source_control_id") or doc.get("control_id") or control_id,
        "question": doc.get("question") or doc.get("problem"),
        "answer": doc.get("answer") or doc.get("solution"),
        "reasoning": doc.get("reasoning"),
        "review_status": doc.get("review_status"),
        "source": collection,
    }


def _guess_nodes_from_entities(
    *,
    client,
    text: str,
    resolved: list[dict],
    external: list[dict],
    unresolved: list[dict],
) -> dict:
    """Return did-you-mean/closest candidates without making them evidence.

    `glossary` is the grounded/unsupported contract. These nodes are only
    reviewer affordances for repair and related-QRA inspection.
    """
    control_nodes: list[dict] = []
    qra_nodes: list[dict] = []
    seen_controls: set[str] = set()
    seen_qras: set[str] = set()

    def add_control_doc(doc: dict | None, *, query: str, match_basis: str, reason: str) -> None:
        if not doc:
            return
        cid = str(doc.get("control_id") or doc.get("id") or doc.get("_id") or "").strip()
        if not cid or cid in seen_controls:
            return
        seen_controls.add(cid)
        control_nodes.append(_candidate_control_node(doc, query=query, match_basis=match_basis, reason=reason))

    resolved_control_ids = [
        str(ent.get("control_id") or ent.get("canonical_id") or "").strip()
        for ent in resolved
        if str(ent.get("control_id") or ent.get("canonical_id") or "").strip()
    ]

    for ent in external + unresolved:
        mention = str(ent.get("mention") or ent.get("term") or ent.get("normalized_text") or "").strip()
        cid = str(ent.get("claimed_against_control_id") or ent.get("control_id") or "").strip()
        if cid and cid not in resolved_control_ids:
            add_control_doc(
                _lookup_sparta_control_doc(client, control_id=cid),
                query=mention or cid,
                match_basis="parenthetical_control_id_context",
                reason=str(ent.get("routing_effect") or ent.get("reason") or "candidate for unsupported term"),
            )
        if mention:
            add_control_doc(
                _lookup_sparta_control_doc_by_name(client, mention),
                query=mention,
                match_basis="exact_control_name",
                reason="did-you-mean control name candidate",
            )

    qra_fields = [
        "_id",
        "_key",
        "qra_id",
        "question",
        "problem",
        "answer",
        "solution",
        "reasoning",
        "control_id",
        "source_control_id",
        "review_status",
    ]
    for cid in resolved_control_ids[:4]:
        for collection in ("sparta_qra_canonical", "sparta_qra_relationship", "sparta_qra"):
            for field in ("source_control_id", "control_id"):
                for doc in _list_collection_docs(
                    client,
                    collection=collection,
                    filters={field: cid},
                    limit=2,
                    return_fields=qra_fields,
                ):
                    key = str(doc.get("_id") or doc.get("_key") or doc.get("qra_id") or "")
                    if not key or key in seen_qras:
                        continue
                    seen_qras.add(key)
                    qra_nodes.append(_candidate_qra_node(doc, query=text, control_id=cid, collection=collection))
                if qra_nodes:
                    break

    return {
        "status": "candidate_only",
        "proof_role": "none",
        "usable_for": ["suggestions", "related_results", "did_you_mean"],
        "not_usable_for": [
            "grounding",
            "proof",
            "approval",
            "answer_authority",
            "evidence_gate",
        ],
        "requires_user_confirmation": True,
        "controls": control_nodes[:8],
        "qras": qra_nodes[:12],
        "notes": [
            "Candidate nodes are did-you-mean or closest related candidates.",
            "Candidate nodes must not satisfy grounding, approval, or evidence gates.",
        ],
    }


def _descriptor_allowed_terms(ent: dict) -> tuple[dict[str, str], str]:
    """Return exact valid descriptor terms for a resolved control entity."""
    control_id = str(ent.get("control_id") or ent.get("canonical_id") or ent.get("label") or "").strip()
    canonical_name = str(ent.get("name") or ent.get("canonical_name") or "").strip()
    category = str(ent.get("category") or ent.get("source_category") or "").strip()
    allowed: dict[str, str] = {}

    def add(value: str, kind: str) -> None:
        cleaned = str(value or "").strip()
        if cleaned:
            allowed.setdefault(cleaned.lower(), kind)

    add(control_id, "control_id")
    add(canonical_name, "control_name")
    add(category, "control_category")
    aliases = ent.get("aliases")
    if isinstance(aliases, list):
        for alias in aliases:
            add(str(alias), "alias")

    allowed_text = "\n".join(
        str(ent.get(field) or "")
        for field in ("description", "definition", "short_description")
    ).lower()
    return allowed, allowed_text


def _parenthetical_descriptors(text: str, resolved: list[dict]) -> tuple[list[dict], list[dict]]:
    """Classify parenthetical control descriptors as grounded or unsupported.

    Example: "CM0029 (Comms Link)" resolves CM0029, then checks the CM0029
    sparta_controls document. If the parenthetical is present as the canonical
    name, category, alias, or exact phrase in the definition/description for
    that specific document, it becomes a grounded descriptor node. Otherwise it
    is an unsupported term.
    """
    lower_text = text.lower()
    descriptors: list[dict] = []
    mismatches: list[dict] = []
    seen: set[tuple[str, int, int]] = set()

    for ent in resolved:
        control_id = str(ent.get("control_id") or ent.get("canonical_id") or ent.get("label") or "").strip()
        if not control_id:
            continue
        allowed_terms, allowed_text = _descriptor_allowed_terms(ent)

        search_from = 0
        control_id_lower = control_id.lower()
        while True:
            cid_start = lower_text.find(control_id_lower, search_from)
            if cid_start < 0:
                break
            cid_end = cid_start + len(control_id)
            open_index = text.find("(", cid_end)
            if open_index < 0 or open_index - cid_end > 3:
                search_from = cid_end
                continue
            close_index = text.find(")", open_index + 1)
            if close_index < 0:
                search_from = cid_end
                continue
            mention = text[open_index + 1:close_index].strip()
            if not mention:
                search_from = close_index + 1
                continue
            mention_lower = mention.lower()
            span_start = text.find(mention, open_index + 1, close_index)
            span_end = span_start + len(mention)
            key = (mention.lower(), span_start, span_end)
            if span_start >= 0 and key not in seen:
                seen.add(key)
                canonical_name = str(ent.get("name") or ent.get("canonical_name") or control_id).strip()
                category = str(ent.get("category") or ent.get("source_category") or "").strip()
                descriptor_kind = allowed_terms.get(mention_lower)
                if descriptor_kind or mention_lower in allowed_text:
                    matched_fields = [descriptor_kind] if descriptor_kind else ["description"]
                    if descriptor_kind in {"control_category", "alias"}:
                        matched_fields = ["category", "alias"]
                    descriptors.append({
                        "id": _term_node_id(f"{control_id}:{mention}", "descriptor"),
                        "mention": mention,
                        "name": mention,
                        "span": [span_start, span_end],
                        "normalized_text": mention_lower,
                        "source": "sparta_controls_parenthetical_guard",
                        "entity_type": "control_descriptor",
                        "control_id": control_id,
                        "descriptor_kind": (
                            "control_category_or_alias"
                            if descriptor_kind in {"control_category", "alias"}
                            else descriptor_kind or "description_phrase"
                        ),
                        "matched_fields": matched_fields,
                        "canonical_name": canonical_name,
                        "category": category,
                        "framework": ent.get("source_framework") or ent.get("framework"),
                        "grounded": True,
                        "exists": True,
                    })
                else:
                    mismatches.append({
                        "id": _unsupported_term_id(mention),
                        "mention": mention,
                        "span": [span_start, span_end],
                        "normalized_text": mention.lower(),
                        "source": "sparta_controls_parenthetical_guard",
                        "entity_type": "unsupported_control_name",
                        "in_security_corpus": False,
                        "routing_effect": f"does_not_match_{control_id}_canonical_name_{canonical_name}",
                        "claimed_against_control_id": control_id,
                        "expected_control_name": canonical_name,
                    })
            search_from = close_index + 1

    return descriptors, mismatches


def _parenthetical_descriptor_keys(text: str, resolved: list[dict]) -> set[tuple[str, tuple[int, int]]]:
    """Return parenthetical spans bound to already resolved first-class IDs."""
    lower_text = text.lower()
    keys: set[tuple[str, tuple[int, int]]] = set()
    for ent in resolved:
        control_id = str(ent.get("control_id") or ent.get("canonical_id") or ent.get("label") or "").strip()
        if not control_id:
            continue
        search_from = 0
        control_id_lower = control_id.lower()
        while True:
            cid_start = lower_text.find(control_id_lower, search_from)
            if cid_start < 0:
                break
            cid_end = cid_start + len(control_id)
            open_index = text.find("(", cid_end)
            if open_index < 0 or open_index - cid_end > 3:
                search_from = cid_end
                continue
            close_index = text.find(")", open_index + 1)
            if close_index < 0:
                search_from = cid_end
                continue
            mention = text[open_index + 1:close_index].strip()
            span_start = text.find(mention, open_index + 1, close_index) if mention else -1
            if span_start >= 0:
                keys.add((mention.lower(), (span_start, span_start + len(mention))))
            search_from = close_index + 1
    return keys


def _enrich_sparta_control_entities(result: dict, text: str) -> dict:
    """Attach canonical sparta_controls metadata to daemon-resolved entities."""
    enriched = dict(result)
    enriched["text"] = text
    client = _make_memory_client()
    try:
        resolved = [
            dict(ent)
            for ent in enriched.get("resolved_entities", [])
            if isinstance(ent, dict)
        ]
        external = [
            dict(ent)
            for ent in enriched.get("external_entities", [])
            if isinstance(ent, dict)
        ]

        by_control_id: dict[str, dict] = {}
        for ent in resolved:
            cid = str(ent.get("canonical_id") or ent.get("control_id") or "").strip()
            if not cid:
                continue
            doc = _lookup_sparta_control_doc(client, control_id=cid)
            if not doc:
                continue
            normalized = _normalize_sparta_control_doc(doc)
            ent.update(normalized)
            ent["mention"] = ent.get("mention") or cid
            by_control_id[normalized.get("control_id", cid)] = ent

        parenthetical_keys = _parenthetical_descriptor_keys(text, resolved)
        descriptor_entities, mismatch_entities = _parenthetical_descriptors(text, resolved)
        still_external: list[dict] = []
        seen_control_ids = set(by_control_id)
        for ent in external:
            mention = str(ent.get("mention") or ent.get("normalized_text") or "").strip()
            span = ent.get("span")
            span_tuple = tuple(span) if isinstance(span, list) and len(span) == 2 else ()
            if (mention.lower(), span_tuple) in parenthetical_keys:
                continue
            doc = _lookup_sparta_control_doc_by_name(client, mention) if mention else None
            if not doc:
                still_external.append(ent)
                continue
            normalized = _normalize_sparta_control_doc(doc)
            cid = normalized.get("control_id")
            if cid in seen_control_ids:
                continue
            promoted = {
                **normalized,
                "mention": ent.get("mention", mention),
                "span": ent.get("span"),
                "match_type": "control_name",
            }
            resolved.append(promoted)
            seen_control_ids.add(cid)
            by_control_id[cid] = promoted

        for mention in _candidate_control_name_mentions(text):
            doc = _lookup_sparta_control_doc_by_name(client, mention)
            if not doc:
                continue
            normalized = _normalize_sparta_control_doc(doc)
            cid = normalized.get("control_id")
            if not cid or cid in seen_control_ids:
                continue
            promoted = {
                **normalized,
                "mention": mention,
                "span": _span_for(text, mention),
                "match_type": "control_name",
            }
            resolved.append(promoted)
            seen_control_ids.add(cid)
            by_control_id[cid] = promoted

        if mismatch_entities:
            for ent in mismatch_entities:
                key = (
                    str(ent.get("mention") or ent.get("normalized_text") or "").strip().lower(),
                    tuple(ent.get("span") or []),
                )
                replaced = False
                for index, existing in enumerate(still_external):
                    existing_key = (
                        str(existing.get("mention") or existing.get("normalized_text") or "").strip().lower(),
                        tuple(existing.get("span") or []),
                    )
                    if existing_key == key:
                        still_external[index] = {**existing, **ent}
                        replaced = True
                        break
                if not replaced:
                    still_external.append(ent)

        control_metadata = [
            _control_metadata_from_entity(ent)
            for ent in resolved
            if ent.get("control_id") or ent.get("canonical_id")
        ]
        existing_metadata = [
            item
            for item in enriched.get("control_metadata", [])
            if isinstance(item, dict)
        ]
        metadata_by_id: dict[str, dict] = {}
        for item in existing_metadata + control_metadata:
            cid = str(item.get("control_id") or "").strip()
            if cid:
                metadata_by_id[cid] = item

        control_ids = [
            str(ent.get("control_id") or ent.get("canonical_id") or "").strip()
            for ent in resolved
        ]
        control_ids = [cid for cid in control_ids if cid]
        existing_control_ids = [
            str(cid).strip()
            for cid in enriched.get("control_ids", [])
            if str(cid).strip()
        ]
        all_control_ids = list(dict.fromkeys(existing_control_ids + control_ids))

        enriched["resolved_entities"] = resolved
        enriched["external_entities"] = still_external
        enriched["control_metadata"] = list(metadata_by_id.values())
        enriched["control_ids"] = all_control_ids
        enriched["all_control_ids"] = list(dict.fromkeys(
            [str(cid).strip() for cid in enriched.get("all_control_ids", []) if str(cid).strip()]
            + all_control_ids
        ))
        summary = dict(enriched.get("summary", {}))
        summary["resolved_count"] = len(resolved)
        summary["external_count"] = len(still_external)
        summary["control_metadata_count"] = len(enriched["control_metadata"])
        domain_terms = [
            dict(term)
            for term in enriched.get("domain_terms", [])
            if isinstance(term, dict)
        ]
        domain_terms, suppressed_nodes = _split_domain_terms_by_descriptor_coverage(
            domain_terms=domain_terms,
            descriptors=descriptor_entities + mismatch_entities,
        )
        enriched["domain_terms"] = domain_terms
        summary["domain_term_count"] = len(domain_terms)
        enriched["glossary"] = _glossary_from_entities(
            resolved=resolved,
            descriptors=descriptor_entities,
            external=still_external,
            unresolved=[
                dict(ent)
                for ent in enriched.get("unresolved_entities", [])
                if isinstance(ent, dict)
            ],
        )
        enriched["entity_nodes"] = _entity_nodes_from_extraction(
            resolved=resolved,
            descriptors=descriptor_entities,
            external=still_external,
            unresolved=[
                dict(ent)
                for ent in enriched.get("unresolved_entities", [])
                if isinstance(ent, dict)
            ],
            domain_terms=domain_terms,
        )
        enriched["candidate_nodes"] = _guess_nodes_from_entities(
            client=client,
            text=text,
            resolved=resolved,
            external=still_external,
            unresolved=[
                dict(ent)
                for ent in enriched.get("unresolved_entities", [])
                if isinstance(ent, dict)
            ],
        )
        primary_entity_id = enriched["all_control_ids"][0] if enriched["all_control_ids"] else None
        enriched["proof_packet"] = _proof_packet_from_nodes(
            query_text=text,
            nodes=enriched["entity_nodes"],
            suppressed_nodes=suppressed_nodes,
            primary_entity_id=primary_entity_id,
        )
        enriched["display_summary"] = _display_summary_from_proof_packet(enriched["proof_packet"])
        enriched["agent_contract"] = _agent_contract()
        summary["glossary_count"] = len(enriched["glossary"])
        summary["entity_node_count"] = len(enriched["entity_nodes"])
        summary["suppressed_node_count"] = len(suppressed_nodes)
        summary["candidate_control_count"] = len(enriched["candidate_nodes"].get("controls", []))
        summary["candidate_qra_count"] = len(enriched["candidate_nodes"].get("qras", []))
        enriched["summary"] = summary

        if resolved and not still_external and not enriched.get("unresolved_entities"):
            enriched["grounding_ok"] = True
            enriched["technique_coherence"] = {
                "ok": True,
                "detail": "single_entity" if len(enriched["all_control_ids"]) <= 1 else "multiple_entities",
            }
            decision = dict(enriched.get("agent_decision", {}))
            decision["safe_to_answer"] = True
            decision["needs_clarification"] = False
            decision["needs_retry"] = False
            decision["suggested_action"] = None
            decision["reason"] = None
            enriched["agent_decision"] = decision

        return enriched
    finally:
        client.close()


def _add_command(commands: list[dict], *, text: str, mention: str, command: str, kind: str) -> None:
    normalized = mention.lower()
    for existing in commands:
        if str(existing.get("mention", "")).lower() == normalized and existing.get("command") == command:
            return
    commands.append({
        "mention": mention,
        "command": command,
        "kind": kind,
        "span": _span_for(text, mention),
    })


def _scan_prefixed_commands(text: str) -> list[dict]:
    commands: list[dict] = []
    command_chars = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_")
    for index, char in enumerate(text):
        if char not in {"$", "/"}:
            continue
        if char == "/" and index > 0 and not text[index - 1].isspace():
            continue
        end = index + 1
        while end < len(text) and text[end] in command_chars:
            end += 1
        if end == index + 1:
            continue
        mention = text[index:end]
        command = mention[1:]
        kind = "skill_command" if command in _KNOWN_INTERFACE_SKILLS else "interface_command"
        commands.append({
            "mention": mention,
            "command": command,
            "kind": kind,
            "span": [index, end],
        })
    return commands


def _extract_dataset_references(text: str) -> list[dict]:
    references: list[dict] = []
    cleaned_tokens = [token.strip(".,;:!?()[]{}<>\"'") for token in text.split()]
    lower_tokens = [token.lower() for token in cleaned_tokens]

    for index, token in enumerate(lower_tokens):
        if token not in {"dataset", "datasets"}:
            continue
        if index + 1 >= len(cleaned_tokens):
            continue
        candidate = cleaned_tokens[index + 1]
        if "/" not in candidate:
            continue
        source = "huggingface" if index > 0 and lower_tokens[index - 1] in {"huggingface", "hf"} else "dataset"
        mention_tokens = cleaned_tokens[max(0, index - 1): index + 2] if source == "huggingface" else cleaned_tokens[index:index + 2]
        mention = " ".join(mention_tokens)
        references.append({
            "mention": mention,
            "source": source,
            "dataset_id": candidate,
            "span": _span_for(text, mention),
        })

    return references


def _extract_file_references(text: str) -> list[dict]:
    references: list[dict] = []
    for token in text.split():
        cleaned = token.strip(".,;:()[]{}<>\"'")
        if (
            cleaned.startswith(("./", "../", "~/", "/"))
            or cleaned.lower().endswith(_DATA_FILE_SUFFIXES)
        ):
            references.append({
                "mention": cleaned,
                "kind": "file_reference",
                "span": _span_for(text, cleaned),
            })
    return references


def _enrich_request_cues(result: dict, text: str) -> dict:
    """Add deterministic request-routing cues to daemon extraction output.

    The memory daemon owns grounded entity extraction. This local enrichment keeps
    interface commands and render/data requests explicit for project agents
    without changing the daemon contract or paying model latency.
    """
    enriched = dict(result)
    lower_text = text.lower()
    commands = _scan_prefixed_commands(text)

    phrase_commands = [
        ("d3 graph", "d3-graph", "render_command"),
        ("d3 chart", "d3-chart", "render_command"),
        ("family tree", "family-tree", "data_shape"),
        ("show as artifact", "open-artifact-pane", "ui_command"),
        ("as an artifact", "open-artifact-pane", "ui_command"),
        ("artifact pane", "open-artifact-pane", "ui_command"),
        ("expand", "expand-artifact", "ui_command"),
        ("render table", "render-table", "render_command"),
        ("table", "render-table", "render_command"),
        ("graph", "render-graph", "render_command"),
        ("chart", "render-chart", "render_command"),
        ("analytics describe", "analytics-describe", "interface_command"),
    ]
    for mention, command, kind in phrase_commands:
        if mention in lower_text:
            _add_command(commands, text=text, mention=mention, command=command, kind=kind)

    skill_mentions = sorted({
        str(command.get("command"))
        for command in commands
        if command.get("kind") == "skill_command"
    })
    dataset_references = _extract_dataset_references(text)
    file_references = _extract_file_references(text)
    has_sample_permission = any(term in lower_text for term in _SAMPLE_DATA_TERMS)
    wants_figure = (
        "create-figure" in skill_mentions
        or "figure" in lower_text
        or any(command.get("kind") == "render_command" for command in commands)
    )
    wants_d3 = "d3" in lower_text
    wants_artifact_pane = any(command.get("command") in {"open-artifact-pane", "expand-artifact"} for command in commands)
    wants_analytics = "analytics" in skill_mentions or "analytics describe" in lower_text
    requires_data = wants_figure or wants_analytics or bool(dataset_references or file_references)

    enriched["interface_commands"] = commands
    enriched["skill_mentions"] = skill_mentions
    enriched["dataset_references"] = dataset_references
    enriched["file_references"] = file_references
    enriched["sample_data_allowed"] = has_sample_permission
    enriched["request_cues"] = {
        "wants_figure": wants_figure,
        "wants_d3": wants_d3,
        "wants_artifact_pane": wants_artifact_pane,
        "wants_analytics": wants_analytics,
        "requires_data": requires_data,
        "has_explicit_sample_permission": has_sample_permission,
    }
    return enriched


def _enrich_extraction_result(result: dict, text: str) -> dict:
    result = _enrich_sparta_control_entities(result, text)
    return _enrich_request_cues(result, text)


# (collection, limit, field names) -> (KeywordProcessor, entity_map)
_ENTITY_DICT_CACHE: dict = {}

# Same key -> how many keywords the dictionary actually held. Lets a caller tell
# "no matches" apart from "nothing was loaded", which are otherwise identical.
_LAST_DICTIONARY_SIZE: dict = {}


def dictionary_size(collection: str, limit: int, name_field: str = "name",
                    label_field: str = "control_id", type_field: str = "node_type",
                    framework_field: str = "source_framework") -> int | None:
    """Keywords loaded for the last extraction over this dictionary, or None.

    Exists so an empty result is interpretable: zero matches against 391,227
    keywords is an answer, zero against 798 is a load failure.
    """
    return _LAST_DICTIONARY_SIZE.get(
        (collection, limit, name_field, label_field, type_field, framework_field)
    )


def _extract_nlp_entities(
    *,
    text: str,
    client,
    collection: str,
    name_field: str,
    label_field: str,
    type_field: str,
    framework_field: str,
    scope: str,
    limit: int,
) -> list[dict]:
    from flashtext import KeywordProcessor

    # Dictionary cache, keyed by the fields that determine its contents.
    #
    # The dictionary was refetched and the trie rebuilt on EVERY call: 391k
    # entities means 782 paged requests and ~43s, discarded immediately. Callers
    # linking a corpus of questions paid that per question, which made the only
    # affordable option a small `limit` -- and a small limit returns nothing. At
    # limit=500 a question naming CWE-89, T1190 and Brute Force grounded ZERO
    # entities; uncapped it grounds four, two of them matched by NAME rather than
    # id, which is the whole point of Flashtext over pattern matching.
    #
    # Aho-Corasick handles millions of keywords; the cost was never the matching.
    # Cached per process, so a batch run pays the load once.
    cache_key = (collection, limit, name_field, label_field, type_field, framework_field)
    # The cache guards the LOAD, never the result. An empty extraction is a valid
    # result: keying the cache on "did this text match anything" meant every
    # question that grounded nothing -- most of them in a batch -- fell through
    # and refetched all 391k entities, costing 11s each while the log claimed a
    # cache hit.
    cached = _ENTITY_DICT_CACHE.get(cache_key)

    # 1. Load entity dictionary from ArangoDB collection (skipped when cached)
    try:
        docs = [] if cached is not None else _list_documents_paged(
            client=client,
            collection=collection,
            limit=limit,
            return_fields=[
                "_id",
                "_key",
                name_field,
                label_field,
                type_field,
                framework_field,
                "description",
                "definition",
                "short_description",
                "category",
                "source_category",
                "aliases",
                "cluster",
                "namespace",
            ],
        )
    except Exception as e:
        logger.warning(f"Failed to load entities from {collection}: {e}")
        docs = []

    # 2. Build FlashText keyword processor from entity names, or reuse the
    # cached trie. Rebuilding 391k keywords per call costs ~0.9s on top of the
    # refetch, for a dictionary that has not changed.
    if cached is not None:
        kp, entity_map = cached
    else:
        kp = KeywordProcessor(case_sensitive=False)
        entity_map: dict[str, dict] = {}  # keyword -> entity doc

    for doc in docs:
        name = doc.get(name_field, "")
        label = doc.get(label_field, "")
        etype = doc.get(type_field, "")
        framework = doc.get(framework_field, "")
        doc_id = doc.get("_id", "")
        category = doc.get("category") or doc.get("source_category") or ""
        aliases = doc.get("aliases") if isinstance(doc.get("aliases"), list) else []

        # Add both name and label as keywords (skip short/generic ones)
        for keyword in [name, label, category, *aliases]:
            if keyword and len(keyword) > 3 and keyword not in entity_map:
                kp.add_keyword(keyword, keyword)
                entity_map[keyword] = {
                    "id": doc_id,
                    "key": doc.get("_key", ""),
                    "name": name,
                    "label": label,
                    "type": etype,
                    "framework": framework,
                    "description": doc.get("description") or "",
                    "definition": doc.get("definition") or doc.get("description") or "",
                    "short_description": doc.get("short_description") or (doc.get("description") or "")[:280],
                    "category": category,
                    "aliases": aliases,
                    "cluster": doc.get("cluster", ""),
                    "namespace": doc.get("namespace", ""),
                    "exists": True,
                }

    if cached is None:
        logger.info(f"Loaded {len(entity_map)} entities from {collection} into FlashText")
        _ENTITY_DICT_CACHE[cache_key] = (kp, entity_map)
    else:
        logger.debug(f"Reusing cached {len(entity_map)}-entity dictionary for {collection}")

    # A zero result is ambiguous: either the text genuinely contains no known
    # entity, or the dictionary never loaded. Those look identical to a caller
    # and are indistinguishable in a log that only reports matches. At
    # limit=500 against a 391k collection, a question naming CWE-89 and T1190
    # returned zero -- a bug that read exactly like a correct empty answer.
    #
    # So the dictionary size is always reported. An empty extraction over a
    # 391,227-entity dictionary is an answer; an empty extraction over 798 is a
    # symptom.
    _LAST_DICTIONARY_SIZE[cache_key] = len(entity_map)
    if not entity_map:
        logger.warning(
            f"Entity dictionary for {collection} is EMPTY -- every extraction will "
            "return nothing. This is a load failure, not an absence of matches."
        )

    # 3. Extract mentioned entities from input text
    found_keywords = kp.extract_keywords(text, span_info=True)
    entities = []
    seen_ids = set()
    for keyword, start, end in found_keywords:
        entity = entity_map.get(keyword)
        if entity and entity["id"] not in seen_ids:
            entities.append({**entity, "span": [start, end], "mention": text[start:end]})
            seen_ids.add(entity["id"])

    # 3b. RapidFuzz fallback: if FlashText found nothing, try fuzzy matching
    _fuzzy_enabled = os.environ.get("EXTRACT_ENTITIES_FUZZY", "1").strip().lower() not in {"0", "false", "no"}
    if len(entities) == 0 and _fuzzy_enabled:
        from rapidfuzz import fuzz, process as rfprocess

        # Extract candidate words from text (skip stop words)
        stop = {
            "the", "and", "show", "me", "all", "its", "click", "on", "related",
            "nodes", "what", "how", "does", "is", "are", "to", "of", "in", "a", "an",
        }
        words = [w for w in text.lower().replace(",", " ").replace(".", " ").split() if w not in stop and len(w) > 2]

        # Try each word against entity names.
        #
        # Stage order is deliberate and unchanged: Flashtext (Aho-Corasick) runs
        # FIRST for exact matching, and this fuzzy pass only runs when it found
        # nothing. That ordering is the point -- fuzzy exists to recover what
        # exact matching MISSED, such as "CWE-7" for "CWE-79", so scoring only
        # Flashtext's own hits would find nothing new and make the stage useless.
        #
        # The cost is O(words x dictionary): on a 391k-entity dictionary that
        # measured 11.6s for one question, fine interactively and ruinous across a
        # batch where most items match nothing exactly. Bulk callers set
        # EXTRACT_ENTITIES_FUZZY=0 to keep exact matching and skip did-you-mean.
        all_names = list(entity_map.keys())
        for word in words[:12]:
            matches = rfprocess.extract(word, all_names, scorer=fuzz.ratio, limit=3,
                                        score_cutoff=70)
            for match_name, score, _ in matches:
                entity = entity_map.get(match_name)
                if entity and entity["id"] not in seen_ids:
                    entities.append({
                        **entity,
                        "span": [0, 0],
                        "mention": word,
                        "fuzzy_score": score,
                        "fuzzy_match": match_name,
                    })
                    seen_ids.add(entity["id"])
                    break  # One fuzzy match per word
        if entities:
            logger.info(f"RapidFuzz fallback found {len(entities)} entities")

    # 4. Enrich with /recall context for each entity
    for entity in entities:
        try:
            recall_q = f"{entity['name']} {entity['type']} {entity.get('cluster', '')}"
            r = client.post("/recall", json={
                "q": recall_q,
                "k": 2,
                **({"scope": scope} if scope else {}),
            })
            items = r.json().get("items", [])
            if items:
                entity["recall_context"] = items[0].get("solution", items[0].get("problem", ""))[:200]
        except Exception:
            pass

    return entities


def _extract_sparta_controls_via_daemon(text: str, client) -> list[dict]:
    """Use Memory's canonical SPARTA extractor and hydrate control docs."""
    try:
        daemon = _call_extract_entities(text, include_taxonomy=False, view="verbose")
    except Exception as exc:
        logger.warning(f"daemon /extract-entities failed for sparta_controls: {exc}")
        return []

    resolved = daemon.get("resolved_entities", []) if isinstance(daemon, dict) else []
    entities: list[dict] = []
    seen_control_ids: set[str] = set()
    for ent in resolved:
        if not isinstance(ent, dict):
            continue
        cid = str(ent.get("canonical_id") or "").strip()
        if not cid or cid in seen_control_ids:
            continue
        doc = _lookup_sparta_control_doc(client, control_id=cid)
        entity = _normalize_sparta_control_doc(doc) if doc else {
            "id": cid,
            "canonical_id": cid,
            "control_id": cid,
            "name": ent.get("canonical_name") or cid,
            "canonical_name": ent.get("canonical_name") or cid,
            "label": cid,
            "framework": ent.get("framework") or "",
            "source_framework": ent.get("framework") or "",
            "type": ent.get("entity_type") or "control",
            "entity_type": ent.get("entity_type") or "control",
            "exists": True,
        }
        entity["span"] = ent.get("span")
        entity["mention"] = ent.get("mention")
        entity["match_type"] = ent.get("resolution_method") or "daemon_extract_entities"
        entities.append(entity)
        seen_control_ids.add(cid)

    for mention in _candidate_control_name_mentions(text):
        doc = _lookup_sparta_control_doc_by_name(client, mention)
        if not doc:
            continue
        entity = _normalize_sparta_control_doc(doc)
        cid = str(entity.get("control_id") or "").strip()
        if not cid or cid in seen_control_ids:
            continue
        entity["span"] = _span_for(text, mention)
        entity["mention"] = mention
        entity["match_type"] = "control_name"
        entities.append(entity)
        seen_control_ids.add(cid)

    return entities


def _lookup_token_by_filter(
    *,
    client,
    collection: str,
    token: str,
    name_field: str,
    label_field: str,
    type_field: str,
    framework_field: str,
    limit: int,
) -> dict | None:
    """Look up a single token by exact filter match (control_id or name/label fallback).

    First tries ``filters: {"control_id": token}`` — the canonical lookup for
    sparta_controls. Falls back to a ``q``-based search with name/label exact
    match for other collections.
    """
    # Primary: exact filter on control_id (canonical for sparta_controls)
    try:
        r = client.post("/list", json={
            "collection": collection,
            "limit": 1,
            "filters": {"control_id": token},
            "return_fields": _SPARTA_CONTROL_RETURN_FIELDS,
        })
        docs = r.json().get("documents", [])
        if docs:
            doc = docs[0]
            return _normalize_sparta_control_doc(doc)
    except Exception as e:
        logger.warning(f"filter lookup failed for '{token}' in {collection}: {e}")

    # Fallback: q-based search with name/label exact match
    try:
        r = client.post("/list", json={
            "collection": collection,
            "q": token,
            "limit": limit,
            "return_fields": _SPARTA_CONTROL_RETURN_FIELDS,
        })
        docs = r.json().get("documents", [])
    except Exception as e:
        logger.warning(f"Failed to lookup token '{token}' in {collection}: {e}")
        docs = []

    token_lc = token.lower()
    doc = None
    for candidate in docs:
        name_val = str(candidate.get(name_field, "")).strip()
        label_val = str(candidate.get(label_field, "")).strip()
        if name_val.lower() == token_lc or label_val.lower() == token_lc:
            doc = candidate
            break
    if not doc and docs:
        doc = docs[0]

    if doc:
        if collection == "sparta_controls":
            return _normalize_sparta_control_doc(doc)
        return {
            "id": doc.get("_id", ""),
            "name": doc.get(name_field, ""),
            "label": doc.get(label_field, ""),
            "framework": doc.get(framework_field, ""),
            "type": doc.get(type_field, ""),
            "exists": True,
        }
    return None


def _extract_delimited_entities(
    *,
    text: str,
    client,
    collection: str,
    name_field: str,
    label_field: str,
    type_field: str,
    framework_field: str,
    delimiter: str,
    limit: int,
) -> list[dict]:
    tokens = _split_tokens(text, delimiter)
    entities = []

    for token in tokens:
        hit = _lookup_token_by_filter(
            client=client,
            collection=collection,
            token=token,
            name_field=name_field,
            label_field=label_field,
            type_field=type_field,
            framework_field=framework_field,
            limit=limit,
        )
        if hit:
            entities.append({"token": token, **hit})
        else:
            entities.append({
                "token": token,
                "id": "",
                "name": token,
                "label": "",
                "framework": "",
                "type": "",
                "exists": False,
            })

    return entities


def _call_extract_entities(text: str, include_taxonomy: bool = False, view: str = "legacy") -> dict:
    """Call daemon /extract-entities endpoint via httpx. Returns JSON dict."""
    client = _make_memory_client()
    try:
        r = client.post("/extract-entities", json={
            "text": text,
            "include_taxonomy": include_taxonomy,
            "view": view,
        })
        r.raise_for_status()
        return r.json()
    finally:
        client.close()


@app.command()
def extract(
    text: str = typer.Argument(..., help="Question or claim text to extract entities from"),
    taxonomy: bool = typer.Option(False, "--taxonomy", "-t", help="Include taxonomy bridge attributes"),
    json_output: bool = typer.Option(False, "--json", help="Output JSON"),
    view: str = typer.Option("agent", "--view", help="JSON view: agent, legacy, verbose, or debug"),
    verbose: bool = typer.Option(False, "--verbose", help="Return verbose diagnostic JSON"),
    quiet: bool = typer.Option(False, "--quiet", "-q", help="Suppress Rich formatting"),
) -> None:
    """Extract entities from question text."""
    output_view = "verbose" if verbose else view
    if output_view not in {"agent", "legacy", "verbose", "debug"}:
        raise typer.BadParameter("view must be one of: agent, legacy, verbose, debug")
    daemon_result = _call_extract_entities(text, include_taxonomy=taxonomy, view="verbose")
    result = _enrich_extraction_result(daemon_result, text)
    for authoritative_key in ("proof_packet", "display_summary", "agent_contract"):
        if daemon_result.get(authoritative_key):
            result[authoritative_key] = daemon_result[authoritative_key]

    if json_output or quiet:
        print(json.dumps(_shape_extraction_result(result, output_view), indent=2, default=str))
        return

    # Rich formatted output
    console.print(f"\n[bold]Entity Extraction[/]")
    console.print(f"[dim]Question:[/] {text[:100]}")
    console.print()

    control_ids = result.get("control_ids", [])
    phrase_controls = result.get("phrase_controls", [])
    phrases = result.get("phrases", [])
    all_control_ids = result.get("all_control_ids", [])
    control_metadata = result.get("control_metadata", [])
    related_pairs = result.get("related_pairs", [])
    taxonomy_tags = result.get("taxonomy_tags", {})

    if control_ids:
        console.print(f"[cyan]Regex control IDs:[/] {', '.join(control_ids)}")
    if phrase_controls:
        console.print(f"[cyan]Phrase-discovered:[/] {', '.join(phrase_controls)}")
    if phrases:
        console.print(f"[cyan]Domain phrases:[/] {', '.join(phrases)}")

    console.print(f"\n[bold]All control IDs ({len(all_control_ids)}):[/] {', '.join(all_control_ids)}")

    if control_metadata:
        console.print()
        table = Table(title="Control Metadata")
        table.add_column("Control ID", style="cyan")
        table.add_column("Name")
        table.add_column("Framework", style="green")
        table.add_column("Domain")
        for c in control_metadata:
            table.add_row(
                c.get("control_id", ""),
                str(c.get("name", ""))[:40],
                c.get("framework", ""),
                c.get("domain", ""),
            )
        console.print(table)

    if related_pairs:
        console.print()
        console.print(f"[bold green]Related pairs ({len(related_pairs)}):[/]")
        for pair in related_pairs[:10]:
            console.print(f"  {pair['source']} --[{pair.get('method', '?')}]--> {pair['target']}")
    elif len(all_control_ids) > 1:
        console.print()
        console.print("[bold red]No relationship edges found between controls[/]")
        console.print("[dim]These controls may not be in the same SPARTA technique[/]")

    if taxonomy_tags:
        console.print()
        console.print(f"[bold]Taxonomy tags:[/]")
        for collection, tags in taxonomy_tags.items():
            if tags:
                console.print(f"  {collection}: {', '.join(tags)}")

    # Summary verdict
    console.print()
    if all_control_ids and related_pairs:
        console.print("[bold green]Coherent:[/] Entities share technique relationships")
    elif all_control_ids and not related_pairs and len(all_control_ids) > 1:
        console.print("[bold yellow]Incoherent:[/] Entities exist but no shared technique — may need decomposition")
    elif all_control_ids:
        console.print("[bold green]Single entity:[/] Coherent by definition")
    else:
        console.print("[bold red]No entities found:[/] Question may be off-topic or too vague")


@app.command()
def check(
    text: str = typer.Argument(..., help="Quick coherence check — are all entities in the same technique?"),
) -> None:
    """Quick boolean: are extracted entities coherent (same technique)?"""
    result = _call_extract_entities(text)

    all_control_ids = result.get("all_control_ids", [])
    related_pairs = result.get("related_pairs", [])

    coherent = (
        len(all_control_ids) <= 1
        or len(related_pairs) > 0
    )

    output = {
        "coherent": coherent,
        "control_count": len(all_control_ids),
        "control_ids": all_control_ids,
        "related_pairs": len(related_pairs),
        "phrases": result.get("phrases", []),
    }

    if not coherent:
        output["reason"] = f"Controls {all_control_ids} have no shared relationship edges"

    print(json.dumps(output, indent=2))


@app.command()
def resolve(
    text: str = typer.Argument(..., help="Text to extract entities from"),
    collection: str = typer.Option("sparta_controls", "--collection", "-c", help="ArangoDB collection to match against"),
    name_field: str = typer.Option("name", "--name-field", help="Field containing entity names"),
    label_field: str = typer.Option("control_id", "--label-field", help="Field containing display labels"),
    type_field: str = typer.Option("control_type", "--type-field", help="Field containing entity type"),
    framework_field: str = typer.Option("source_framework", "--framework-field", help="Field containing framework"),
    delimiter: str = typer.Option("nlp", "--delimiter", "-d", help="Entity extraction mode: nlp, auto, or custom delimiter character(s)"),
    scope: str = typer.Option("", "--scope", "-s", help="Scope filter for /recall"),
    json_output: bool = typer.Option(True, "--json", help="Output JSON"),
    limit: int = typer.Option(500, "--limit", help="Max entities to load for FlashText dictionary"),
) -> None:
    """Collection-agnostic entity extraction via FlashText + /memory recall.

    Loads entity names from any ArangoDB collection, builds a FlashText keyword
    processor, extracts mentions from the input text, then enriches with /recall
    context. Zero LLM cost.

    Used by all ux-lab chat wells for entity grounding before intent routing.
    """
    client = _make_memory_client()

    if delimiter.strip().lower() == "nlp":
        # SPARTA fast path: use daemon /extract-entities contract directly.
        # This yields grounded control spans and avoids local vocabulary drift.
        if collection == "sparta_controls":
            entities = _extract_sparta_controls_via_daemon(text, client)
            if not entities:
                entities = _extract_nlp_entities(
                    text=text,
                    client=client,
                    collection=collection,
                    name_field=name_field,
                    label_field=label_field,
                    type_field=type_field,
                    framework_field=framework_field,
                    scope=scope,
                    limit=limit,
                )
        else:
            entities = _extract_nlp_entities(
                text=text,
                client=client,
                collection=collection,
                name_field=name_field,
                label_field=label_field,
                type_field=type_field,
                framework_field=framework_field,
                scope=scope,
                limit=limit,
            )
    else:
        entities = _extract_delimited_entities(
            text=text,
            client=client,
            collection=collection,
            name_field=name_field,
            label_field=label_field,
            type_field=type_field,
            framework_field=framework_field,
            delimiter=delimiter,
            limit=limit,
        )

    # 5. Output
    result = {
        "text": text,
        "collection": collection,
        "delimiter": delimiter,
        "entity_count": len(entities),
        "entities": entities,
        "entity_names": [e.get("name", "") for e in entities],
        "entity_ids": [e.get("id", "") for e in entities if e.get("id")],
    }

    client.close()
    print(json.dumps(result, indent=2, default=str))


def _run_default_mode() -> None:
    """Default stdin mode: invoked without a subcommand.

    Reads text from stdin, then:
    - If ``--delimiter`` is present: split tokens and lookup each via /list filters.
    - Otherwise: use the NLP extraction path via daemon /extract-entities.

    Flags:
      --delimiter <mode>    auto | custom string | omit for NLP
      --collection <name>   ArangoDB collection (default: sparta_controls)
      --name-field <field>  name field (default: name)
      --label-field <field> label field (default: control_id)
      --framework-field <f> framework field (default: source_framework)
      --type-field <field>  type field (default: node_type)
      --limit <n>           max docs for NLP dictionary (default: 500)
      --scope <s>           scope filter for /recall (NLP mode only)
    """
    # Use typer-style argument parsing for the default stdin mode
    default_app = typer.Typer()

    @default_app.command()
    def _default_cmd(
        delimiter: str = typer.Option(None, "--delimiter", "-d", help="Entity extraction mode: auto, custom string, or omit for NLP"),
        collection: str = typer.Option("sparta_controls", "--collection", "-c", help="ArangoDB collection"),
        name_field: str = typer.Option("name", "--name-field", help="Field containing entity names"),
        label_field: str = typer.Option("control_id", "--label-field", help="Field containing display labels"),
        framework_field: str = typer.Option("source_framework", "--framework-field", help="Field containing framework"),
        type_field: str = typer.Option("control_type", "--type-field", help="Field containing entity type"),
        limit: int = typer.Option(500, "--limit", help="Max entities to load for FlashText dictionary"),
        scope: str = typer.Option("", "--scope", help="Scope filter for /recall"),
    ) -> None:
        text = sys.stdin.read().strip()
        if not text:
            print(json.dumps({"error": "no input text", "entities": [], "entity_count": 0}))
            return

        client = _make_memory_client()

        if delimiter is not None:
            entities = _extract_delimited_entities(
                text=text,
                client=client,
                collection=collection,
                name_field=name_field,
                label_field=label_field,
                type_field=type_field,
                framework_field=framework_field,
                delimiter=delimiter,
                limit=limit,
            )
        else:
            if collection == "sparta_controls":
                entities = _extract_sparta_controls_via_daemon(text, client)
                if not entities:
                    entities = _extract_nlp_entities(
                        text=text,
                        client=client,
                        collection=collection,
                        name_field=name_field,
                        label_field=label_field,
                        type_field=type_field,
                        framework_field=framework_field,
                        scope=scope,
                        limit=limit,
                    )
            else:
                entities = _extract_nlp_entities(
                    text=text,
                    client=client,
                    collection=collection,
                    name_field=name_field,
                    label_field=label_field,
                    type_field=type_field,
                    framework_field=framework_field,
                    scope=scope,
                    limit=limit,
                )

        result = {
            "text": text,
            "collection": collection,
            "delimiter": delimiter,
            "entity_count": len(entities),
            "entities": entities,
            "entity_names": [e.get("name", "") for e in entities],
            "entity_ids": [e.get("id", "") for e in entities if e.get("id")],
        }
        client.close()
        print(json.dumps(result, indent=2, default=str))

    default_app()


# Known Typer subcommands — anything else (or no args besides flags) goes to default stdin mode.
_TYPER_SUBCOMMANDS = {"extract", "check", "resolve"}


if __name__ == "__main__":
    # If the first non-flag argument is a known subcommand, delegate to Typer.
    # Otherwise, run the default stdin mode.
    positional_args = [a for a in sys.argv[1:] if not a.startswith("-")]
    if positional_args and positional_args[0] in _TYPER_SUBCOMMANDS:
        app()
    else:
        _run_default_mode()
