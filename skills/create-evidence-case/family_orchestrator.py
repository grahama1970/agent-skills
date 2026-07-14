"""Requirement-first F36 family evidence orchestration.

This module keeps the canonical engineering QRA immutable while resolving
candidate SPARTA relationships from the accepted local corpus.  Every result is
quarantined for human review; no result is promoted to reviewed truth here.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import time
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

from runner import EvidenceCaseRunner


MEMORY_API_BASE = os.environ.get("MEMORY_API_BASE", "http://127.0.0.1:8601")
MEMORY_TIMEOUT = httpx.Timeout(60.0, connect=5.0)
ORCHESTRATOR_CONTRACT_ID = "f36-create-evidence-case-family-v1.3"
VARIANT_PLAN = (
    ("operator", "simple"),
    ("project_manager", "simple"),
    ("systems_engineer", "medium"),
    ("cybersecurity_compliance_officer", "medium"),
    ("mission_assurance_cybersecurity_reviewer", "advanced"),
)
MATERIAL_REQUIREMENT_KEYS = (
    "requirement_id",
    "title",
    "statement",
    "rationale",
    "major_system_id",
    "subsystem_id",
    "component_family_id",
    "requirement_type",
    "verification_method",
    "verification_artifact_types",
    "lifecycle_phase_ids",
)
DISALLOWED_EDGE_STATES = {"rejected", "superseded", "stale"}
TRIAGE_DISPOSITIONS = {
    "candidate_applicable",
    "adjacent",
    "informational",
    "candidate_non_applicable",
    "clarify",
}
STOPWORDS = {
    "and", "are", "for", "from", "how", "shall", "that", "the", "this",
    "what", "when", "where", "which", "with", "must", "required", "outcome",
    "condition", "behavior", "results", "synthetic", "function", "plan",
}


def canonical_bytes(value: Any) -> bytes:
    return json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=True).encode()


def sha256_bytes(value: bytes) -> str:
    return f"sha256:{hashlib.sha256(value).hexdigest()}"


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def requirement_content_hash(requirement: dict[str, Any]) -> str:
    material = {key: requirement.get(key) for key in MATERIAL_REQUIREMENT_KEYS}
    return sha256_bytes(canonical_bytes(material))


def _revision_id(requirement: dict[str, Any]) -> str:
    return str(
        (((requirement.get("traceability") or {}).get("supersession") or {}).get("revision_id"))
        or ""
    )


def validate_family_input(envelope: dict[str, Any]) -> list[str]:
    issues: list[str] = []
    if envelope.get("schema") != "f36.evidence_orchestration_family_input.v1":
        issues.append("invalid input schema")
    requirement = envelope.get("requirement") or {}
    family = envelope.get("family") or {}
    corpus = envelope.get("sparta_corpus") or {}
    triage = envelope.get("triage") or {"disposition": "candidate_applicable"}

    if not requirement.get("requirement_id"):
        issues.append("missing requirement_id")
    if family.get("requirement_id") != requirement.get("requirement_id"):
        issues.append("family requirement_id mismatch")
    if family.get("requirement_revision_id") != _revision_id(requirement):
        issues.append("family requirement_revision_id mismatch")
    if family.get("requirement_content_hash") != requirement_content_hash(requirement):
        issues.append("family requirement_content_hash mismatch")
    if family.get("primary_component_family_id") != requirement.get("component_family_id"):
        issues.append("family component owner mismatch")
    if not family.get("canonical_question") or not family.get("canonical_answer"):
        issues.append("canonical question or answer missing")
    if triage.get("disposition") not in TRIAGE_DISPOSITIONS:
        issues.append("invalid triage disposition")

    variants = family.get("variants") or []
    if len(variants) != 5:
        issues.append("family must contain exactly five variants")
    expected_ids = [f"{family.get('engineering_qra_family_id')}-V{i:02d}" for i in range(1, 6)]
    if [v.get("variant_id") for v in variants] != expected_ids:
        issues.append("variant IDs or order mismatch")
    if [(v.get("role"), v.get("difficulty")) for v in variants] != list(VARIANT_PLAN):
        issues.append("variant role/difficulty order mismatch")
    if any(v.get("answer") != family.get("canonical_answer") for v in variants):
        issues.append("variant answer differs from canonical answer")
    intent_hashes = {v.get("intent_hash") for v in variants}
    if len(intent_hashes) != 1 or None in intent_hashes:
        issues.append("variant intent hashes are not identical and nonempty")

    source_path = Path(str(corpus.get("source_path") or ""))
    if corpus.get("release_id") != "sparta-excel-v3.1-9cbd7eef12547bd0":
        issues.append("unexpected SPARTA release ID")
    if not source_path.is_file():
        issues.append("SPARTA source workbook missing")
    elif corpus.get("release_hash") != sha256_file(source_path):
        issues.append("SPARTA source workbook hash mismatch")
    return issues


def _post(client: httpx.Client, path: str, payload: dict[str, Any]) -> dict[str, Any]:
    response = client.post(path, json=payload)
    response.raise_for_status()
    value = response.json()
    if not isinstance(value, dict):
        raise ValueError(f"{path} returned non-object JSON")
    return value


def _exact_documents(
    client: httpx.Client,
    collection: str,
    filters: dict[str, Any],
    limit: int = 100,
) -> list[dict[str, Any]]:
    result = _post(
        client,
        "/list",
        {"collection": collection, "filters": filters, "limit": limit, "offset": 0},
    )
    return [item for item in result.get("documents") or [] if isinstance(item, dict)]


def _exact_node(client: httpx.Client, control_id: str) -> dict[str, Any] | None:
    docs = _exact_documents(client, "sparta_controls", {"control_id": control_id}, limit=20)
    exact = [doc for doc in docs if doc.get("control_id") == control_id]
    if len(exact) != 1:
        return None
    doc = exact[0]
    return {
        "persisted_node_id": doc.get("_id"),
        "control_id": control_id,
        "name": doc.get("name"),
        "description": doc.get("description"),
        "source_framework": doc.get("source_framework"),
        "source_version": doc.get("source_version"),
        "status": doc.get("status"),
    }


def _usable_edge(doc: dict[str, Any], source_id: str, target_id: str) -> bool:
    return bool(
        doc.get("_id")
        and doc.get("source_control_id") == source_id
        and doc.get("target_control_id") == target_id
        and doc.get("edge_type")
        and doc.get("type")
        and doc.get("review_status") not in DISALLOWED_EDGE_STATES
        and not doc.get("normal_coverage_excluded", False)
        and doc.get("relationship_integrity_status") not in {
            "edge_type_unknown_quarantined",
            "rejected_orphan_endpoint",
        }
    )


def _split_cell(value: Any) -> set[str]:
    if value is None:
        return set()
    return {part.strip() for part in str(value).split(",") if part.strip()}


def _tokens(value: Any) -> set[str]:
    tokens: set[str] = set()
    for token in re.findall(r"[a-z0-9]+", str(value or "").lower()):
        if len(token) <= 2 or token in STOPWORDS:
            continue
        for suffix in ("ing", "ied", "ed", "es", "s"):
            if token.endswith(suffix) and len(token) - len(suffix) >= 4:
                token = token[: -len(suffix)]
                break
        tokens.add(token)
    return tokens


def _workbook_ranked_ids(
    workbook_path: Path, query: str, limit: int = 300
) -> dict[str, tuple[int, int]]:
    """Rank exact workbook entities by lexical overlap with the family obligation."""
    query_tokens = _tokens(query)
    scores: list[tuple[int, str]] = []
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for sheet_name, text_columns in (
            ("SPARTA Techniques", ("Name", "Description")),
            ("SPARTA Countermeasures", ("Name", "Description")),
            ("NIST References with SSG", ("Name", "Discussion", "Control")),
        ):
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "") for value in next(rows)]
            id_index = headers.index("ID")
            text_indexes = [headers.index(column) for column in text_columns if column in headers]
            for row in rows:
                control_id = str(row[id_index] or "").strip()
                if not control_id:
                    continue
                text = " ".join(str(row[index] or "") for index in text_indexes)
                score = len(query_tokens & _tokens(text))
                if score:
                    scores.append((score, control_id))
    finally:
        workbook.close()
    scores.sort(key=lambda item: (-item[0], item[1]))
    return {
        control_id: (score, rank)
        for rank, (score, control_id) in enumerate(scores[:limit], start=1)
    }


def _workbook_relation(
    workbook_path: Path,
    source_id: str,
    target_id: str,
    source_framework: str,
    target_framework: str,
) -> dict[str, Any] | None:
    """Prove supported directed relationships in the pinned v3.1 workbook."""
    source_fw = source_framework.lower()
    target_fw = target_framework.lower()
    if {source_fw, target_fw} != {"nist", "sparta"}:
        return None

    nist_id = source_id if source_fw == "nist" else target_id
    sparta_id = target_id if target_fw == "sparta" else source_id
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["SPARTA Techniques"]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(rows)]
        id_index = headers.index("ID")
        nist_index = headers.index("NIST Rev5 Controls")
        for row_number, row in enumerate(rows, start=2):
            if str(row[id_index] or "").strip() != sparta_id:
                continue
            if nist_id in _split_cell(row[nist_index]):
                return {
                    "workbook_sheet": sheet.title,
                    "workbook_row": row_number,
                    "relationship_column": "NIST Rev5 Controls",
                    "sparta_control_id": sparta_id,
                    "related_control_id": nist_id,
                    "direction": f"{source_id}->{target_id}",
                }
        return None
    finally:
        workbook.close()


def _candidate_chains(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for rank, item in enumerate(items, start=1):
        evidence_case = item.get("evidence_case") or {}
        for chain in evidence_case.get("crosswalk_chains") or []:
            if not isinstance(chain, dict):
                continue
            source_id = str(chain.get("source_id") or "")
            target_id = str(chain.get("target_id") or "")
            if not source_id or not target_id or (source_id, target_id) in seen:
                continue
            seen.add((source_id, target_id))
            candidates.append(
                {
                    "recall_rank": rank,
                    "source_qra_key": item.get("_key"),
                    "source_qra_control_id": item.get("control_id"),
                    "source_qra_question": item.get("question"),
                    "source_id": source_id,
                    "target_id": target_id,
                    "source_framework": str(chain.get("source_framework") or ""),
                    "target_framework": str(chain.get("target_framework") or ""),
                    "candidate_method": chain.get("method"),
                }
            )
    return candidates


def _resolve_paths(
    client: httpx.Client,
    items: list[dict[str, Any]],
    workbook_path: Path,
    release_id: str,
    release_hash: str,
    query: str,
    max_paths: int = 3,
) -> tuple[list[dict[str, Any]], list[str]]:
    proofs: list[dict[str, Any]] = []
    grounded_ids: set[str] = set()
    ranked_ids = _workbook_ranked_ids(workbook_path, query)
    candidates = _candidate_chains(items)
    candidates.sort(
        key=lambda candidate: (
            -sum(ranked_ids.get(candidate[key], (0, 10_000))[0] for key in ("source_id", "target_id")),
            max(ranked_ids.get(candidate[key], (0, 10_000))[1] for key in ("source_id", "target_id")),
            candidate["recall_rank"],
            candidate["source_id"],
            candidate["target_id"],
        )
    )
    for candidate in candidates:
        if (
            candidate["source_id"] not in ranked_ids
            or candidate["target_id"] not in ranked_ids
        ):
            continue
        source_id = candidate["source_id"]
        target_id = candidate["target_id"]
        edge_docs = _exact_documents(
            client,
            "sparta_relationships",
            {"source_control_id": source_id, "target_control_id": target_id},
            limit=20,
        )
        for edge in edge_docs:
            if not _usable_edge(edge, source_id, target_id):
                continue
            source_framework = str(edge.get("source_framework") or candidate["source_framework"])
            target_framework = str(edge.get("target_framework") or candidate["target_framework"])
            corpus_proof = _workbook_relation(
                workbook_path,
                source_id,
                target_id,
                source_framework,
                target_framework,
            )
            if not corpus_proof:
                continue
            source_node = _exact_node(client, source_id)
            target_node = _exact_node(client, target_id)
            if not source_node or not target_node:
                continue
            grounded_ids.update((source_id, target_id))
            proof = {
                "path_signature": sha256_bytes(
                    canonical_bytes([edge.get("_id"), source_id, target_id, release_hash])
                ),
                "source_qra_key": candidate["source_qra_key"],
                "source_qra_control_id": candidate["source_qra_control_id"],
                "recall_rank": candidate["recall_rank"],
                "workbook_relevance_score": (
                    ranked_ids[source_id][0] + ranked_ids[target_id][0]
                ),
                "workbook_relevance_ranks": {
                    source_id: ranked_ids[source_id][1],
                    target_id: ranked_ids[target_id][1],
                },
                "nodes": [source_node, target_node],
                "persisted_edge_ids": [edge.get("_id")],
                "edges": [
                    {
                        "persisted_edge_id": edge.get("_id"),
                        "source_id": source_id,
                        "target_id": target_id,
                        "source_framework": source_framework,
                        "target_framework": target_framework,
                        "relationship_type": edge.get("edge_type"),
                        "stored_type": edge.get("type"),
                        "method": edge.get("method"),
                        "review_status": edge.get("review_status"),
                        "direction": f"{source_id}->{target_id}",
                    }
                ],
                "corpus_release": {
                    "release_id": release_id,
                    "release_hash": release_hash,
                    **corpus_proof,
                },
                "authority_state": "persisted_candidate_pending_review",
            }
            proofs.append(proof)
            break
        if len(proofs) >= max_paths:
            break
    return proofs, sorted(grounded_ids)


def _event(stage: str, state: str, detail: dict[str, Any]) -> dict[str, Any]:
    return {"stage": stage, "state": state, "timestamp_ns": time.time_ns(), "detail": detail}


def _assertion_role(route: str) -> tuple[str, bool]:
    if route == "organizational_assurance":
        return "assurance_context", False
    return "mitigation", True


def _build_assertion_manifest(
    family: dict[str, Any],
    route: str,
    paths: list[dict[str, Any]],
) -> dict[str, Any]:
    role, material = _assertion_role(route)
    assertions: list[dict[str, Any]] = []
    for path in paths:
        source, target = path["nodes"]
        assertion_id = "F36B-MAP-" + hashlib.sha256(
            canonical_bytes(
                [family["engineering_qra_family_id"], path["path_signature"], role, material]
            )
        ).hexdigest()[:20]
        assertions.append(
            {
                "assertion_id": assertion_id,
                "relationship_role": role,
                "candidate_material": material,
                "source_entity": {
                    "framework": source["source_framework"],
                    "id": source["control_id"],
                    "name": source["name"],
                },
                "target_entity": {
                    "framework": target["source_framework"],
                    "id": target["control_id"],
                    "name": target["name"],
                },
                "path_proof_id": path["path_signature"],
                "persisted_edge_ids": path["persisted_edge_ids"],
            }
        )
    manifest = {
        "schema": "f36.family_candidate_assertion_manifest.v1",
        "family_id": family["engineering_qra_family_id"],
        "assertions": assertions,
    }
    manifest["manifest_hash"] = sha256_bytes(canonical_bytes(manifest))
    return manifest


def _assertion_question(intent: dict[str, Any], assertion: dict[str, Any]) -> str:
    source = assertion["source_entity"]
    target = assertion["target_entity"]
    return (
        f"For an engineering obligation to {intent.get('required_behavior') or 'perform the declared behavior'} "
        f"under {intent.get('condition') or 'the declared condition'} for "
        f"{intent.get('protected_object_or_interface') or 'the declared protected object or interface'}, "
        f"with the intended outcome {intent.get('expected_outcome') or 'the declared expected outcome'}, "
        f"what evidence supports or contradicts treating the relationship between "
        f"{source['framework']} {source['id']} ({source['name']}) and "
        f"{target['framework']} {target['id']} ({target['name']}) as a "
        f"{assertion['relationship_role']} mapping?"
    )


def run_family_orchestration(
    envelope: dict[str, Any],
    output_path: Path,
    *,
    live: bool = True,
) -> dict[str, Any]:
    issues = validate_family_input(envelope)
    if issues:
        raise ValueError("; ".join(issues))

    requirement = envelope["requirement"]
    family = envelope["family"]
    corpus = envelope["sparta_corpus"]
    canonical_answer_hash = sha256_bytes(family["canonical_answer"].encode())
    input_fingerprint = sha256_bytes(
        canonical_bytes({"orchestrator_contract_id": ORCHESTRATOR_CONTRACT_ID, "input": envelope})
    )
    events: list[dict[str, Any]] = []
    events.append(_event("input_validation", "passed", {"input_fingerprint": input_fingerprint}))

    decomposition = {
        "engineering_obligation_id": family.get("engineering_obligation_id"),
        **(family.get("canonical_intent") or {}),
    }
    events.append(_event("decompose", "passed", {"obligation_count": 1}))

    intent = family.get("canonical_intent") or {}
    query = " ".join(
        part
        for part in (
            requirement.get("statement"),
            intent.get("protected_object_or_interface"),
            intent.get("condition"),
            intent.get("required_behavior"),
        )
        if part
    )

    triage = envelope.get("triage") or {"disposition": "candidate_applicable"}
    triage_disposition = triage["disposition"]
    events.append(_event("applicability_triage", "complete", triage))
    recall_result: dict[str, Any] = {"found": False, "items": []}
    recall_items: list[dict[str, Any]] = []
    path_proofs: list[dict[str, Any]] = []
    grounded_ids: list[str] = []
    if triage_disposition in {"candidate_applicable", "adjacent"}:
        with httpx.Client(base_url=MEMORY_API_BASE, timeout=MEMORY_TIMEOUT) as client:
            recall_result = _post(
                client,
                "/recall",
                {"q": query, "collections": ["sparta_qra"], "k": 12, "threshold": 0.3},
            )
            recall_items = [item for item in recall_result.get("items") or [] if isinstance(item, dict)]
            events.append(
                _event(
                    "sparta_qra_recall",
                    "passed" if recall_items else "empty",
                    {"item_count": len(recall_items), "found": bool(recall_result.get("found"))},
                )
            )
            path_proofs, grounded_ids = _resolve_paths(
                client,
                recall_items,
                Path(corpus["source_path"]),
                corpus["release_id"],
                corpus["release_hash"],
                query,
            )
    else:
        events.append(
            _event(
                "sparta_qra_recall",
                "skipped_by_triage",
                {"triage_disposition": triage_disposition},
            )
        )

    events.append(
        _event(
            "exact_grounding_and_path_resolution",
            "passed" if path_proofs else "no_path",
            {"grounded_control_ids": grounded_ids, "path_count": len(path_proofs)},
        )
    )

    route = envelope.get("applicability_route", "unresolved")
    assertion_manifest = _build_assertion_manifest(family, route, path_proofs)
    assertion_outcomes: list[dict[str, Any]] = []
    execution_live = live and triage_disposition in {"candidate_applicable", "adjacent"}
    if execution_live:
        runner = EvidenceCaseRunner()
        for assertion in assertion_manifest["assertions"]:
            evidence_query = _assertion_question(intent, assertion)
            runner_result = runner.run(
                claim_text=evidence_query,
                category="compliance",
                show_progress=False,
                expected_control_ids=[
                    assertion["source_entity"]["id"],
                    assertion["target_entity"]["id"],
                ],
                exact_path_proven=True,
            )
            runner_state = str(
                (runner_result.get("verdict") or {}).get("state") or "inconclusive"
            )
            evidence_verdict = {
                "satisfied": "SATISFIED",
                "not_satisfied": "NOT_SATISFIED",
                "inconclusive": "INCONCLUSIVE",
                "ambiguous_entity": "NOT_SATISFIED",
            }.get(runner_state, "INCONCLUSIVE")
            outcome = {
                **assertion,
                "evidence_query": evidence_query,
                "evidence_verdict": evidence_verdict,
                "evidence_case_id": (runner_result.get("claim") or {}).get("id"),
                "persistence": runner_result.get("persistence"),
                "gate_trace": runner_result.get("gate_trace") or [],
                "raw_result_hash": sha256_bytes(canonical_bytes(runner_result)),
                "raw_result": runner_result,
            }
            assertion_outcomes.append(outcome)
            events.append(
                _event(
                    "assertion_evidence_evaluation",
                    "complete",
                    {
                        "assertion_id": assertion["assertion_id"],
                        "evidence_verdict": evidence_verdict,
                    },
                )
            )

    evidence_required = triage_disposition in {"candidate_applicable", "adjacent"}
    sparta_applicability = {
        "candidate_applicable": "sparta_applicable",
        "adjacent": "sparta_candidate_adjacent",
        "informational": "sparta_applicable",
        "candidate_non_applicable": "not_sparta_applicable",
        "clarify": "unresolved",
    }[triage_disposition]
    if execution_live:
        evidence_execution_state = "completed"
        crosswalk_resolution_state = "exact_path" if path_proofs else "no_exact_path"
    else:
        evidence_execution_state = (
            "not_required"
            if triage_disposition in {"informational", "candidate_non_applicable"}
            else "not_run"
        )
        evidence_verdict = None
        crosswalk_resolution_state = "not_attempted"

    if triage_disposition == "candidate_non_applicable":
        terminal_state = "NOT_APPLICABLE_PENDING_REVIEW"
        family_disposition = "candidate_non_applicable"
    elif triage_disposition == "informational":
        terminal_state = "INFORMATIONAL"
        family_disposition = "informational"
    elif triage_disposition == "clarify":
        terminal_state = "CLARIFY"
        family_disposition = "clarify"
    elif not path_proofs:
        terminal_state = "INCONCLUSIVE"
        evidence_verdict = "INCONCLUSIVE"
        family_disposition = "unresolved_evidence"
    else:
        material_outcomes = [row for row in assertion_outcomes if row["candidate_material"]]
        if not material_outcomes:
            terminal_state = "INFORMATIONAL"
            evidence_verdict = None
            family_disposition = "informational"
        elif any(row["evidence_verdict"] == "NOT_SATISFIED" for row in material_outcomes):
            terminal_state = "NOT_SATISFIED"
            evidence_verdict = "NOT_SATISFIED"
            family_disposition = "material_negative"
        elif any(row["evidence_verdict"] != "SATISFIED" for row in material_outcomes):
            terminal_state = "INCONCLUSIVE"
            evidence_verdict = "INCONCLUSIVE"
            family_disposition = "unresolved_evidence"
        else:
            terminal_state = "SATISFIED"
            evidence_verdict = "SATISFIED"
            family_disposition = "satisfied_candidate"
    events.append(_event("verdict", "complete", {"terminal_state": terminal_state}))

    family_snapshot_id = "F36B-ECF-" + hashlib.sha256(
        f"{family['engineering_qra_family_id']}|{input_fingerprint}".encode()
    ).hexdigest()[:20]
    snapshot = {
        "schema": "f36.family_evidence_case_snapshot.v1",
        "family_evidence_case_snapshot_id": family_snapshot_id,
        "orchestrator_contract_id": ORCHESTRATOR_CONTRACT_ID,
        "input_fingerprint": input_fingerprint,
        "mocked": False,
        "live": execution_live,
        "synthetic": True,
        "operational_authority": False,
        "accepted": False,
        "review_state": "pending",
        "quarantine_state": "quarantined_pending_human_review",
        "requirement": {
            "requirement_id": requirement["requirement_id"],
            "requirement_revision_id": _revision_id(requirement),
            "requirement_content_hash": requirement_content_hash(requirement),
            "primary_component_family_id": requirement["component_family_id"],
        },
        "engineering_qra_family": {
            "engineering_qra_family_id": family["engineering_qra_family_id"],
            "canonical_question": family["canonical_question"],
            "canonical_answer": family["canonical_answer"],
            "canonical_answer_hash": canonical_answer_hash,
            "canonical_intent": family.get("canonical_intent"),
            "variant_count": len(family["variants"]),
            "variant_evidence_runs": 0,
            "variant_inheritance": [
                {
                    "variant_id": variant["variant_id"],
                    "intent_hash": variant["intent_hash"],
                    "answer_hash": sha256_bytes(variant["answer"].encode()),
                    "inherited_family_snapshot_id": family_snapshot_id,
                }
                for variant in family["variants"]
            ],
        },
        "applicability": {
            "route": envelope.get("applicability_route", "unresolved"),
            "state": triage_disposition,
            "sparta_applicability": sparta_applicability,
            "reason_codes": triage.get("reason_codes") or [],
            "review_state": "pending",
        },
        "evidence_required": evidence_required,
        "evidence_execution_state": evidence_execution_state,
        "crosswalk_resolution_state": crosswalk_resolution_state,
        "evidence_verdict": evidence_verdict,
        "family_disposition": family_disposition,
        "family_orchestration_runs": 1,
        "assertion_evidence_evaluations": len(assertion_outcomes),
        "stage_receipts": {
            "decomposition": decomposition,
            "sparta_qra_recall": {
                "query_hash": sha256_bytes(query.encode()),
                "found": bool(recall_result.get("found")),
                "item_count": len(recall_items),
                "item_keys": [item.get("_key") for item in recall_items],
                "raw_content_hash": sha256_bytes(canonical_bytes(recall_result)),
            },
            "entity_grounding": {
                "resolved_control_ids": grounded_ids,
                "invented_control_ids": [],
                "exact_node_count": len(grounded_ids),
            },
            "coherence": {
                "state": "candidate_coherent_group" if path_proofs else "unresolved",
                "basis": "same recalled sparta_qra item and exact persisted chain",
            },
            "path_resolution": {
                "sparta_release_id": corpus["release_id"],
                "sparta_release_hash": corpus["release_hash"],
                "path_proofs": path_proofs,
            },
            "assertion_manifest": assertion_manifest,
            "formal_proof": {"state": "not_run"},
            "assertion_evidence_outcomes": assertion_outcomes,
        },
        "terminal_disposition": terminal_state,
        "binding_registry_state": "not_promoted",
        "projection_eligibility": {
            "candidate_review_mode": bool(path_proofs),
            "reviewed_default_mode": False,
            "posture_grounding_numerator": False,
            "supply_chain_sparta_overlay": False,
        },
        "stage_events": events,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    events_path = output_path.with_suffix(".events.jsonl")
    events_path.write_text("".join(json.dumps(event, sort_keys=True) + "\n" for event in events))
    snapshot["artifact_paths"] = {
        "snapshot": str(output_path),
        "events": str(events_path),
    }
    snapshot["snapshot_content_hash"] = sha256_bytes(canonical_bytes(snapshot))
    output_path.write_text(json.dumps(snapshot, indent=2, ensure_ascii=True) + "\n")
    return snapshot
