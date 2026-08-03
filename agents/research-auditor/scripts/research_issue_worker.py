#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path
from typing import Any, Mapping, Sequence

import httpx

COMMON_DIR = Path(__file__).resolve().parents[2] / "qra-auditor" / "scripts"
if str(COMMON_DIR) not in sys.path:
    sys.path.insert(0, str(COMMON_DIR))

from lane_worker_common import (
    build_tau_agent_handoff,
    claim_one,
    run_registry_decision,
    update_issue,
    utc_now,
    write_json,
    write_tau_handoff_artifacts,
)

DEFAULT_QUEUE = Path("/mnt/storage12tb/media/agents/shared/monitor-sparta/repair_queue.jsonl")
DEFAULT_RUN_ROOT = Path("/mnt/storage12tb/skills/review-db/outputs/research-auditor")
DEFAULT_MEMORY_ROOT = Path("/home/graham/workspace/experiments/memory")
DEFAULT_MEMORY_BASE_URL = "http://127.0.0.1:8601"
DEFAULT_SPARTA_ROOT = Path("/home/graham/workspace/experiments/sparta")
DEFAULT_AGENT_SKILLS_ROOT = Path("/home/graham/workspace/experiments/agent-skills")
MITRE_MOBILE_ATTACK_JSON = "https://raw.githubusercontent.com/mitre/cti/master/mobile-attack/mobile-attack.json"
ALLOWED_LANES = {"source_fetch", "source_research", "source_truth_lookup", "source_url_fetch_evidence"}


def issue_next_action(issue: Mapping[str, Any]) -> dict[str, Any]:
    slice_ = issue.get("slice") if isinstance(issue.get("slice"), Mapping) else {}
    next_action = slice_.get("next_action") if isinstance(slice_.get("next_action"), Mapping) else {}
    return dict(next_action)


def source_manifest_path(issue: Mapping[str, Any]) -> Path | None:
    next_action = issue_next_action(issue)
    candidates = [
        next_action.get("manifest_path"),
        next_action.get("source_text_qra_manifest"),
        issue.get("source_text_qra_manifest"),
    ]
    slice_ = issue.get("slice") if isinstance(issue.get("slice"), Mapping) else {}
    candidates.append(slice_.get("source_text_qra_manifest"))
    for value in candidates:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return Path(text)
    return None


def memory_list(collection: str, filters: Mapping[str, Any], *, limit: int = 10) -> list[dict[str, Any]]:
    with httpx.Client(base_url=DEFAULT_MEMORY_BASE_URL, timeout=httpx.Timeout(20.0, connect=2.0)) as client:
        resp = client.post("/list", json={"collection": collection, "filters": dict(filters), "limit": limit})
        resp.raise_for_status()
        payload = resp.json()
    docs = payload.get("documents") if isinstance(payload, Mapping) else []
    return [dict(doc) for doc in docs if isinstance(doc, Mapping)]


def memory_query(aql: str, bind_vars: Mapping[str, Any]) -> list[dict[str, Any]]:
    with httpx.Client(base_url=DEFAULT_MEMORY_BASE_URL, timeout=httpx.Timeout(20.0, connect=2.0)) as client:
        resp = client.post("/query", json={"aql": aql, "bind_vars": dict(bind_vars)})
        resp.raise_for_status()
        payload = resp.json()
    docs = payload.get("documents") if isinstance(payload, Mapping) else []
    return [dict(doc) for doc in docs if isinstance(doc, Mapping)]


def load_mobile_attack_by_external_id() -> dict[str, dict[str, Any]]:
    try:
        resp = httpx.get(MITRE_MOBILE_ATTACK_JSON, timeout=httpx.Timeout(20.0, connect=2.0), follow_redirects=True)
        resp.raise_for_status()
        payload = resp.json()
    except Exception as exc:
        return {"__error__": {"error": f"{type(exc).__name__}: {exc}", "source_url": MITRE_MOBILE_ATTACK_JSON}}
    out: dict[str, dict[str, Any]] = {}
    for obj in payload.get("objects", []):
        if not isinstance(obj, Mapping):
            continue
        for ref in obj.get("external_references") or []:
            if isinstance(ref, Mapping) and ref.get("external_id"):
                out[str(ref["external_id"])] = dict(obj)
    return out


def workbook_row_by_id(control_id: str) -> dict[str, Any] | None:
    workbook = DEFAULT_SPARTA_ROOT / "data/source/SPARTA-Data.xlsx"
    if not workbook.exists():
        return None
    try:
        import openpyxl

        wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    except Exception:
        return None
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        except StopIteration:
            continue
        for row in rows:
            values = {
                headers[idx]: row[idx]
                for idx in range(min(len(headers), len(row)))
                if headers[idx] and row[idx] is not None and str(row[idx]).strip()
            }
            if any(str(value).strip() == control_id for value in values.values()):
                return {
                    "source": f"workbook:{workbook}:sheet:{sheet}",
                    "row": {key: str(value) for key, value in values.items()},
                }
    return None


def enrich_source_target(action: Mapping[str, Any], *, mobile_attack: Mapping[str, Any]) -> dict[str, Any]:
    target = dict(action)
    lane = str(action.get("lane") or "")
    if lane == "control_text":
        key = str(action.get("key") or "")
        control_id = str(action.get("control_id") or "")
        docs = memory_list("sparta_controls", {"_key": key}, limit=1) if key else []
        current = docs[0] if docs else {}
        source_framework = str(current.get("source_framework") or action.get("source_framework") or "")
        evidence: dict[str, Any] = {
            "current": {
                "_key": current.get("_key"),
                "control_id": current.get("control_id") or control_id,
                "name": current.get("name"),
                "description": current.get("description"),
                "source_framework": source_framework,
                "control_type": current.get("control_type"),
            }
        }
        if source_framework == "ATT_CK_Mobile" and control_id:
            obj = mobile_attack.get(control_id)
            if isinstance(obj, Mapping):
                evidence["authoritative_source"] = {
                    "source": MITRE_MOBILE_ATTACK_JSON,
                    "external_id": control_id,
                    "name": obj.get("name"),
                    "description": obj.get("description"),
                    "revoked": obj.get("revoked"),
                    "x_mitre_deprecated": obj.get("x_mitre_deprecated"),
                    "modified": obj.get("modified"),
                }
                if obj.get("revoked") is True:
                    target["source_resolution"] = "source_confirms_revoked_or_unusable"
                    target["recommended_next_owner"] = "dba-auditor-v2"
                    target["recommended_next_lane"] = "source_text_status_repair"
                    target["recommended_action"] = "mark_control_deprecated_or_excluded_with_rollback"
                elif len(str(obj.get("description") or "").strip()) > 20:
                    target["source_resolution"] = "source_description_available"
                    target["recommended_next_owner"] = "dba-auditor-v2"
                    target["recommended_next_lane"] = "source_text_status_repair"
                    target["recommended_action"] = "backfill_control_description_with_rollback"
                else:
                    target["source_resolution"] = "source_description_still_stub"
                    target["recommended_next_owner"] = "operator"
                    target["recommended_action"] = "upstream_source_contains_stub"
            else:
                target["source_resolution"] = "authoritative_source_not_found"
        elif source_framework == "ISO" and control_id:
            row = workbook_row_by_id(control_id)
            evidence["authoritative_source"] = row
            target["source_resolution"] = "source_heading_only_no_narrative_available" if row else "authoritative_source_not_found"
            target["recommended_next_owner"] = "dba-auditor-v2"
            target["recommended_next_lane"] = "source_text_status_repair"
            target["recommended_action"] = "mark_heading_only_non_generation_or_excluded_with_rollback"
        else:
            target["source_resolution"] = "unsupported_control_source_framework"
        target["evidence"] = evidence
        return target

    if lane == "url_text":
        key = str(action.get("key") or "")
        url_id = str(action.get("url_id") or "")
        url_docs = memory_list("sparta_urls", {"_key": key}, limit=1) if key else []
        knowledge_docs = (
            memory_query(
                """
FOR d IN sparta_url_knowledge
  FILTER TO_STRING(d.url_id) == @url_id
  RETURN KEEP(d, "_key", "url_id", "text", "topic", "semantic_sync_state", "qdrant_point_id")
""",
                {"url_id": url_id},
            )
            if url_id
            else []
        )
        nonstub_knowledge = [
            doc
            for doc in knowledge_docs
            if len(" ".join(str(doc.get("text") or "").split())) >= 200
        ]
        target["evidence"] = {
            "current_url": url_docs[0] if url_docs else {},
            "knowledge_doc_count": len(knowledge_docs),
            "nonstub_knowledge_doc_count": len(nonstub_knowledge),
            "knowledge_doc_keys": [doc.get("_key") for doc in knowledge_docs],
        }
        target["source_resolution"] = "external_url_fetch_required" if not nonstub_knowledge else "url_knowledge_exists"
        if nonstub_knowledge:
            target["recommended_next_owner"] = "qra-auditor"
            target["recommended_action"] = "resume_qra_manifest_after_source_text_available"
        else:
            target["recommended_next_owner"] = "research-auditor"
            target["recommended_action"] = "fetch_url_and_extract_text_with_rollback_backfill_manifest"
        return target

    target["source_resolution"] = "not_a_source_text_target"
    return target


def build_source_evidence_manifest(issue: Mapping[str, Any], *, run_dir: Path) -> dict[str, Any]:
    manifest_path = source_manifest_path(issue)
    out_path = run_dir / "source_evidence_manifest.json"
    if manifest_path is None:
        payload = {
            "schema": "research_auditor.source_evidence_manifest.v1",
            "terminal_status": "MISSING_SOURCE_MANIFEST",
            "source_manifest": None,
            "source_targets": [],
            "source_target_count": 0,
            "qra_targets_deferred_count": 0,
            "created_at": utc_now(),
        }
        write_json(out_path, payload)
        return payload
    if not manifest_path.exists():
        payload = {
            "schema": "research_auditor.source_evidence_manifest.v1",
            "terminal_status": "SOURCE_MANIFEST_NOT_FOUND",
            "source_manifest": str(manifest_path),
            "source_targets": [],
            "source_target_count": 0,
            "qra_targets_deferred_count": 0,
            "created_at": utc_now(),
        }
        write_json(out_path, payload)
        return payload

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    actions = manifest.get("actions") if isinstance(manifest, Mapping) else []
    if not isinstance(actions, list):
        actions = []
    source_targets = [
        dict(action)
        for action in actions
        if isinstance(action, Mapping) and action.get("lane") in {"control_text", "url_text"}
    ]
    mobile_attack = load_mobile_attack_by_external_id()
    enriched_source_targets = [enrich_source_target(action, mobile_attack=mobile_attack) for action in source_targets]
    resolutions: dict[str, int] = {}
    for target in enriched_source_targets:
        resolution = str(target.get("source_resolution") or "unknown")
        resolutions[resolution] = resolutions.get(resolution, 0) + 1
    qra_targets = [
        dict(action)
        for action in actions
        if isinstance(action, Mapping) and str(action.get("lane") or "").endswith("_qra")
    ]
    payload = {
        "schema": "research_auditor.source_evidence_manifest.v1",
        "terminal_status": "SOURCE_TARGETS_IDENTIFIED" if source_targets else "NO_SOURCE_TARGETS",
        "source_manifest": str(manifest_path),
        "source_manifest_status": manifest.get("status") if isinstance(manifest, Mapping) else None,
        "source_targets": enriched_source_targets,
        "source_target_count": len(enriched_source_targets),
        "source_resolution_counts": resolutions,
        "qra_targets_deferred_count": len(qra_targets),
        "qra_targets_deferred_lanes": sorted({str(action.get("lane")) for action in qra_targets}),
        "mutation_enabled": False,
        "operator_approval_required": bool(source_targets),
        "created_at": utc_now(),
    }
    write_json(out_path, payload)
    return payload


def run_url_fetches(source_evidence: dict[str, Any], *, run_dir: Path) -> dict[str, Any]:
    targets = source_evidence.get("source_targets") if isinstance(source_evidence.get("source_targets"), list) else []
    fetcher = DEFAULT_AGENT_SKILLS_ROOT / "skills" / "fetcher" / "run.sh"
    fetch_results: list[dict[str, Any]] = []
    for target in targets:
        if not isinstance(target, dict):
            continue
        if target.get("source_resolution") != "external_url_fetch_required":
            continue
        evidence = target.get("evidence") if isinstance(target.get("evidence"), Mapping) else {}
        current_url = evidence.get("current_url") if isinstance(evidence.get("current_url"), Mapping) else {}
        url = str(current_url.get("url") or target.get("url") or "").strip()
        url_id = str(target.get("url_id") or current_url.get("url_id") or "").strip()
        if not url or not url_id:
            continue
        fetch_dir = run_dir / f"url_{url_id}_fetch"
        fetch_dir.mkdir(parents=True, exist_ok=True)
        stdout_path = fetch_dir / "fetcher.stdout.log"
        stderr_path = fetch_dir / "fetcher.stderr.log"
        cmd = [str(fetcher), "get", url, "--out", str(fetch_dir)]
        proc = subprocess.run(cmd, text=True, capture_output=True, check=False, timeout=300)
        stdout_path.write_text(proc.stdout or "", encoding="utf-8")
        stderr_path.write_text(proc.stderr or "", encoding="utf-8")
        summary_path = fetch_dir / "consumer_summary.json"
        result = {
            "url_id": url_id,
            "url": url,
            "cmd": cmd,
            "exit_code": proc.returncode,
            "stdout_path": str(stdout_path),
            "stderr_path": str(stderr_path),
            "consumer_summary": str(summary_path) if summary_path.exists() else None,
            "usable_for_dewey_backfill": False,
            "reason": None,
        }
        if proc.returncode == 0 and summary_path.exists():
            try:
                summary = json.loads(summary_path.read_text(encoding="utf-8"))
                item = (summary.get("items") or [{}])[0]
                artifacts = item.get("artifacts") if isinstance(item, Mapping) and isinstance(item.get("artifacts"), Mapping) else {}
                text_path = artifacts.get("extracted_text_path")
                text = Path(str(text_path)).read_text(encoding="utf-8", errors="replace") if text_path else ""
                result["text_length"] = len(" ".join(text.split()))
                result["final_downloaded_url"] = item.get("final_downloaded_url") if isinstance(item, Mapping) else None
                result["usable_for_dewey_backfill"] = int(result["text_length"] or 0) >= 200
                result["reason"] = "fetcher_text_available" if result["usable_for_dewey_backfill"] else "fetcher_text_too_short"
                if result["usable_for_dewey_backfill"]:
                    target["source_resolution"] = "url_fetch_text_available"
                    target["recommended_next_owner"] = "dba-auditor-v2"
                    target["recommended_next_lane"] = "source_url_text_backfill"
                    target["recommended_action"] = "backfill_url_text_with_rollback"
                    target["fetcher_summary"] = str(summary_path)
            except Exception as exc:  # noqa: BLE001 - evidence should capture fetch parse failure
                result["reason"] = f"fetcher_summary_parse_failed:{type(exc).__name__}:{exc}"
        else:
            result["reason"] = "fetcher_failed"
        fetch_results.append(result)
    source_evidence["url_fetch_results"] = fetch_results
    source_evidence["url_fetch_result_count"] = len(fetch_results)
    source_evidence["source_resolution_counts"] = {}
    for target in targets:
        if isinstance(target, Mapping):
            resolution = str(target.get("source_resolution") or "unknown")
            source_evidence["source_resolution_counts"][resolution] = source_evidence["source_resolution_counts"].get(resolution, 0) + 1
    write_json(run_dir / "source_evidence_manifest.json", source_evidence)
    return source_evidence


def run(args: argparse.Namespace) -> tuple[int, dict]:
    run_id = args.run_id or "research-auditor-run"
    run_dir = Path(args.run_root) / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    queue = Path(args.queue)
    issue = claim_one(queue, owner="research-auditor", run_id=run_id, allowed_lanes=ALLOWED_LANES)
    if issue is None:
        receipt = {
            "schema": "research_auditor.issue_worker.receipt.v1",
            "run_id": run_id,
            "run_dir": str(run_dir),
            "queue_path": str(queue),
            "terminal_status": "NO_READY_ISSUE",
            "mocked": False,
            "live": True,
            "created_at": utc_now(),
        }
        write_json(run_dir / "receipt.json", receipt)
        return 3, receipt

    write_json(run_dir / "issue.json", issue)
    issue_id = str(issue.get("issue_id") or "")
    lane = str(issue.get("lane") or "")
    source_evidence = build_source_evidence_manifest(issue, run_dir=run_dir)
    source_evidence = run_url_fetches(source_evidence, run_dir=run_dir)
    source_target_count = int(source_evidence.get("source_target_count") or 0)
    if source_evidence.get("terminal_status") == "NO_SOURCE_TARGETS":
        decision_name = "source_fetch_not_required"
        decision_status = "DONE"
        queue_status = "DONE"
        decision_reason = "no_source_text_targets_in_manifest"
        summary = "Ryan inspected the source/QRA manifest and found no source-fetch targets."
        rationale = "The manifest contains QRA generation/linkage work only; Qbert/Petey remain the owners for QRA generation gates."
    elif source_target_count > 0:
        resolution_counts = source_evidence.get("source_resolution_counts") if isinstance(source_evidence.get("source_resolution_counts"), Mapping) else {}
        unresolved = int(resolution_counts.get("external_url_fetch_required") or 0) + int(resolution_counts.get("authoritative_source_not_found") or 0)
        source_targets = source_evidence.get("source_targets") if isinstance(source_evidence.get("source_targets"), list) else []
        dewey_targets = [
            target
            for target in source_targets
            if isinstance(target, Mapping)
            and target.get("recommended_next_owner") == "dba-auditor-v2"
            and target.get("recommended_next_lane") == "source_text_status_repair"
        ]
        url_backfill_targets = [
            target
            for target in source_targets
            if isinstance(target, Mapping)
            and target.get("recommended_next_owner") == "dba-auditor-v2"
            and target.get("recommended_next_lane") == "source_url_text_backfill"
        ]
        if dewey_targets:
            decision_name = "source_text_status_repair_required"
            decision_status = "NEEDS_AGENT"
            queue_status = "DONE"
            decision_reason = "source_evidence_contains_deterministic_dewey_repairs"
            next_action = {
                "type": "create_queue_issue",
                "owner_subagent": "dba-auditor-v2",
                "owner_display_name": "Dewey",
                "lane": "source_text_status_repair",
                "collection": "sparta_controls",
                "scope": "source_evidence_manifest",
                "limit": len(dewey_targets),
                "mutation_allowed": True,
                "source_evidence_manifest": str(run_dir / "source_evidence_manifest.json"),
                "manifest_path": str(run_dir / "source_evidence_manifest.json"),
                "success_signal": "source_text_status_repair_proof_ok",
                "issue_id": f"{issue_id}:dewey-source-text-status",
                "blocked_issue_id": issue_id,
            }
            needed_agent = "dba-auditor-v2"
            needed_display_name = "Dewey"
        elif url_backfill_targets:
            target = url_backfill_targets[0]
            decision_name = "source_url_text_backfill_required"
            decision_status = "NEEDS_AGENT"
            queue_status = "DONE"
            decision_reason = "url_fetch_evidence_contains_dewey_backfill"
            next_action = {
                "type": "create_queue_issue",
                "owner_subagent": "dba-auditor-v2",
                "owner_display_name": "Dewey",
                "lane": "source_url_text_backfill",
                "collection": "sparta_url_knowledge",
                "scope": "one_url_fetcher_artifact",
                "limit": 1,
                "mutation_allowed": True,
                "url_id": target.get("url_id"),
                "fetcher_summary": target.get("fetcher_summary"),
                "success_signal": "source_url_text_backfill_proof_ok",
                "issue_id": f"{issue_id}:dewey-url-text-{target.get('url_id')}",
                "blocked_issue_id": issue_id,
            }
            needed_agent = "dba-auditor-v2"
            needed_display_name = "Dewey"
        else:
            decision_name = "source_evidence_packet_written"
            decision_status = "NEEDS_HUMAN"
            queue_status = "OPERATOR_REQUIRED"
            decision_reason = "source_evidence_ready_but_apply_or_external_fetch_not_approved"
            next_action = {}
            needed_agent = None
            needed_display_name = None
        summary = (
            f"Ryan wrote source evidence for {source_target_count} source text target(s); "
            f"resolution_counts={dict(resolution_counts)}; "
            f"dewey_handoff_targets={len(dewey_targets) if 'dewey_targets' in locals() else 0}."
        )
        rationale = (
            "Ryan resolved local/upstream source state without mutating the database. "
            f"{unresolved} target(s) still require external URL fetch or human-approved backfill/apply. "
            "Deterministic control source-status repairs are handed to Dewey as a separate one-lane queue issue."
        )
    else:
        decision_name = "source_fetch_manifest_missing"
        decision_status = "NEEDS_HUMAN"
        queue_status = "OPERATOR_REQUIRED"
        decision_reason = str(source_evidence.get("terminal_status") or "source_manifest_unavailable").lower()
        summary = "Ryan could not inspect exact source targets because the source manifest was missing or unavailable."
        rationale = "Source-fetch work must fail closed until monitor-sparta/Qbert supplies an exact review-required manifest."
        next_action = {}
        needed_agent = None
        needed_display_name = None

    if source_evidence.get("terminal_status") == "NO_SOURCE_TARGETS":
        next_action = {}
        needed_agent = None
        needed_display_name = None

    registry = run_registry_decision(
        memory_root=Path(args.memory_root),
        output=run_dir / "registry_decision.json",
        decision_type="source_fetch",
        decision=decision_name,
        status=decision_status,
        issuer_subagent="research-auditor",
        issuer_display_name="Ryan",
        subject_subagent="research-auditor",
        subject_display_name="Ryan",
        lane=lane,
        collection=str(issue.get("collection") or ""),
        issue_id=issue_id,
        summary=summary,
        rationale=rationale,
        decision_reason=decision_reason,
        next_action=next_action,
        needed_agent=needed_agent,
        needed_display_name=needed_display_name,
        run_id=run_id,
        receipt_path=run_dir / "receipt.json",
        artifact_paths={
            "issue": str(run_dir / "issue.json"),
            "source_evidence_manifest": str(run_dir / "source_evidence_manifest.json"),
            "source_manifest": source_evidence.get("source_manifest"),
        },
    )
    if decision_status == "NEEDS_AGENT" and needed_agent:
        tau_next_agent = needed_agent
        tau_reason = f"{needed_display_name or needed_agent} owns the next bounded repair lane."
    elif decision_status == "DONE":
        tau_next_agent = "qra-auditor"
        tau_reason = "Qbert can continue because no source-fetch targets remain for Ryan."
    else:
        tau_next_agent = "human"
        tau_reason = "Human/operator review is required before further source repair routing."
    tau_handoff = write_tau_handoff_artifacts(
        run_dir,
        filename_stem="ryan_source_fetch",
        handoff=build_tau_agent_handoff(
            previous_subagent="research-auditor",
            next_agent=tau_next_agent,
            reason=tau_reason,
            result_status=decision_status,
            result_summary=summary,
            context_summary="monitor-sparta queued a source evidence issue for Ryan.",
            rationale=rationale,
            stop_condition=(
                "Next owner writes a registry decision or queue receipt for the handed-off source issue."
                if tau_next_agent != "human"
                else "Human/operator accepts, rejects, or supplies missing external authority."
            ),
            issue_id=issue_id,
            evidence=[str(run_dir / "issue.json"), str(run_dir / "source_evidence_manifest.json"), str(run_dir / "registry_decision.json")],
            artifacts=[str(run_dir / "source_evidence_manifest.json")],
            required_evidence=[
                "source_evidence_manifest",
                "subagent_approval_registry decision row",
            ],
        ),
    )
    queue_update = update_issue(
        queue,
        issue,
        status=queue_status,
        run_id=run_id,
        event="ryan_source_fetch_triaged",
        fields={
            "registry_decision_key": registry.get("decision_key"),
            "tau_handoff_path": tau_handoff.get("handoff_path"),
            "tau_validation_path": tau_handoff.get("validation_path"),
            "tau_handoff_ok": tau_handoff.get("ok"),
            "blocked_reason": decision_reason if queue_status != "DONE" else None,
            "source_evidence_manifest": str(run_dir / "source_evidence_manifest.json"),
            "source_target_count": source_target_count,
        },
    )
    receipt = {
        "schema": "research_auditor.issue_worker.receipt.v1",
        "run_id": run_id,
        "run_dir": str(run_dir),
        "queue_path": str(queue),
        "claimed_issue_id": issue_id,
        "lane": lane,
        "terminal_status": queue_status,
        "tau_handoff": tau_handoff,
        "source_evidence": source_evidence,
        "registry_decision": registry,
        "queue_update": queue_update,
        "mocked": False,
        "live": True,
        "forbidden_paths": {
            "repair_cycle_invoked": False,
            "health_fix_invoked": False,
            "database_mutation": False,
            "source_truth_invented": False,
        },
    }
    write_json(run_dir / "receipt.json", receipt)
    return 0, receipt


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="cmd", required=True)
    run_p = sub.add_parser("run")
    run_p.add_argument("--run-id")
    run_p.add_argument("--run-root", type=Path, default=DEFAULT_RUN_ROOT)
    run_p.add_argument("--queue", type=Path, default=DEFAULT_QUEUE)
    run_p.add_argument("--memory-root", type=Path, default=DEFAULT_MEMORY_ROOT)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    if args.cmd == "run":
        rc, receipt = run(args)
        print(json.dumps(receipt, indent=2, sort_keys=True))
        return rc
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
